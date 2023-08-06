# -*- coding: UTF-8 -*-
from openpyxl import Workbook, load_workbook
def copy_sheet(src_xlsx, ssheetname, dst_xlsx, nsheetname=None):
    if nsheetname == None:
        nsheetname = ssheetname
    try:
        sw = load_workbook(f'{src_xlsx}')
    except KeyError:
        raise KeyError('旧工作簿不存在 The old xlsx is not exists')

    try:
        dw = load_workbook(f'{dst_xlsx}')
    except FileNotFoundError:
        dw = Workbook()

    try:
        sheet = dw[f'{nsheetname}']
    except KeyError:
        sheet = dw.create_sheet(f'{nsheetname}')

    try:
        src_sheet = sw[f'{ssheetname}']
    except KeyError:
        raise KeyError('源工作簿文件不存在该工作簿 The sheet does not exist in the source file')

    # for i in range(1):
    #     sheet.delete_rows(2000)
    for sheet in dw:
        sheet.delete_rows(1, 3000)

    for row in src_sheet.iter_rows():
        # print(row)
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        sheet.append(row_list)
    dw.save(f'{dst_xlsx}')


