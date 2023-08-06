import typing

from openpyxl.cell import Cell
from returns import returns

from .utils.class_collector import CollectorMeta, ListCollector


class ExcelModel(metaclass=CollectorMeta):
    columns: ListCollector['TColumnDef'] = ListCollector()

    @classmethod
    def as_table(
            cls,
            *,
            table_def_class: typing.Type['TTableDef'] = None,
            **kwargs,
    ) -> 'TTableDef':
        if table_def_class is None:
            from .tables import ExcelTableDefinition
            table_def_class = ExcelTableDefinition
        return table_def_class(cls, **kwargs)

    def __init__(
            self,
            table: 'TTable',
            idx: int,
            row_num: int,
    ):
        self.table = table
        self.idx = idx
        self.row_num = row_num

        self.values_cache = {}

    def __eq__(self, other: typing.Self) -> bool:
        if other is None or not isinstance(other, ExcelModel):
            return False

        return (
                self.table == other.table
                and self.idx == other.idx
                and self.row_num == other.row_num
        )

    def __bool__(self) -> bool:
        for cell in self.cells:
            if cell.value is not None:
                return True
        return False

    @returns(dict)
    def as_dict(self) -> dict[str, typing.Any]:
        for column in self.columns:
            yield column.name, column.__get__(self)

    def set_dict(
            self,
            mapping: typing.Mapping[str, typing.Any] = None,
            /,
            **kwargs,
    ):
        if mapping is not None:
            for k, v in mapping.items():
                setattr(self, k, v)
        for k, v in kwargs.items():
            setattr(self, k, v)

    def cell(self, col_num: int) -> Cell:
        return self.table.cell(self.row_num, col_num)

    def cell0(self, col_idx: int) -> Cell:
        return self.cell(col_idx + 1)

    def cella(self, attr: str) -> Cell:
        return self.cell(getattr(self.table, attr).col_num)

    @property
    def cells(self) -> typing.Sequence[Cell]:
        return self.table.ws[self.row_num]


TModel = typing.TypeVar('TModel', bound=ExcelModel)

if typing.TYPE_CHECKING:
    from .tables import TTableDef, TTable
    from .columns import TColumnDef
