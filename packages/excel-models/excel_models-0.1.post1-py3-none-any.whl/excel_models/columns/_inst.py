import typing
from functools import cached_property

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

from ._base import TColumnDef
from ..tables import TTable


class ExcelColumn(typing.Generic[TTable, TColumnDef]):
    def __init__(
            self,
            table: TTable,
            column_def: TColumnDef,
            col_num: int,
    ):
        self.table = table
        self.column_def = column_def
        self.col_num = col_num

    @cached_property
    def col_letter(self) -> str:
        return get_column_letter(self.col_num)

    def __eq__(self, other: typing.Self) -> bool:
        if other is None or not isinstance(other, ExcelColumn):
            return False

        return (
                self.table == other.table
                and self.column_def == other.column_def
                and self.col_num == other.col_num
        )

    def __getitem__(self, idx: int | slice):
        if isinstance(idx, slice):
            return [
                self.column_def.__get__(row)
                for row in self.table[idx]
            ]

        return self.column_def.__get__(self.table[idx])

    def __iter__(self) -> typing.Iterator:
        for row in self.table:
            yield self.column_def.__get__(row)

    def __setitem__(self, idx: int | slice, value):
        if isinstance(idx, slice):
            for row, v in zip(self.table[idx], value, strict=True):
                self.column_def.__set__(row, v)
            return

        self.column_def.__set__(self.table[idx], value)

    def __delitem__(self, idx: int | slice):
        if isinstance(idx, slice):
            for row in self.table[idx]:
                self.column_def.__delete__(row)
            return

        self.column_def.__delete__(self.table[idx])

    def cell(self, row_num: int) -> Cell:
        return self.table.cell(row_num, self.col_num)

    def cell0(self, idx: int) -> Cell:
        return self.cell(self.table.get_row_num(idx))

    @property
    def cells(self) -> typing.Sequence[Cell]:
        return self.table.ws[self.col_letter][self.table.get_row_num(-1):]


TColumn = typing.TypeVar('TColumn', bound=ExcelColumn)
