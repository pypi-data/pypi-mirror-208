import typing
from functools import cached_property

from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from returns import returns

from ._def import TTableDef
from ..db import TDB
from ..exceptions import ColumnNotFound
from ..models import TModel


class ExcelTable(typing.Generic[TDB, TModel, TTableDef]):
    def __init__(
            self,
            db: TDB,
            table_def: TTableDef,
            ws: Worksheet,
    ):
        self.db = db
        self.table_def = table_def
        self.ws = ws

        self._columns_cache = {}

    def __eq__(self, other: typing.Self) -> bool:
        if other is None or not isinstance(other, ExcelTable):
            return False

        return (
                self.db == other.db
                and self.table_def == other.table_def
                and self.ws == other.ws
        )

    @property
    def model(self) -> typing.Type[TModel]:
        return self.table_def.model

    @property
    def _max_row(self) -> int:
        return self.ws.max_row

    def __len__(self) -> int:
        return self._max_row - self.table_def.title_row

    def get_row_num(self, idx: int) -> int:
        return self.table_def.title_row + idx + 1

    def get_idx(self, row_num: int) -> int:
        return row_num - self.table_def.title_row - 1

    def _get_range(
            self,
            s: slice = None,
            *,
            start=None,
            stop=None,
            step=None,
    ) -> list[int]:
        if s is not None:
            start = s.start
            stop = s.stop
            step = s.step

        return list(range(len(self)))[start:stop:step]

    def __getitem__(self, idx: int | slice) -> typing.Union[TModel, list[TModel]]:
        if isinstance(idx, slice):
            return [
                self[i]
                for i in self._get_range(idx)
            ]
        return self.model(self, idx, self.get_row_num(idx))

    def __iter__(self) -> typing.Iterator[TModel]:
        for i in self._get_range():
            yield self[i]

    def _get_column_def(self, attr: str) -> 'TColumnDef':
        for column in self.model.columns:
            if column.attr == attr:
                return column
        raise AttributeError(attr)

    def _get_col_num(self, name: str) -> int:
        for cell in self.ws[self.table_def.title_row]:
            if cell.value == name:
                return cell.column
        raise ColumnNotFound(name)

    def _get_column(self, attr: str) -> 'TColumn':
        if attr not in self._columns_cache:
            column_def = self._get_column_def(attr)
            col_num = self._get_col_num(column_def.name)
            from ..columns import ExcelColumn
            self._columns_cache[attr] = ExcelColumn(self, column_def, col_num)
        return self._columns_cache[attr]

    @cached_property
    @returns(tuple)
    def columns(self) -> typing.Sequence['TColumn']:
        for column in self.model.columns:
            yield self._get_column(column.attr)

    def __getattr__(self, attr: str) -> 'TColumn':
        return self._get_column(attr)

    def cell(self, row_num, col_num) -> Cell:
        return self.ws.cell(row_num, col_num)

    def new(self) -> TModel:
        self.ws.append([])
        row_num = self.ws._current_row  # noqa: pycharm
        return self[self.get_idx(row_num)]

    @property
    def _max_column_letter(self) -> str:
        return getattr(self, self.model.columns[-1].attr).col_letter

    @property
    def _filter_ref_str(self) -> str:
        return f'A{self.table_def.title_row}:{self._max_column_letter}{self._max_row}'

    def add_filter(self):
        self.ws.auto_filter.ref = self._filter_ref_str


TTable = typing.TypeVar('TTable', bound=ExcelTable)

if typing.TYPE_CHECKING:
    from ..columns import TColumnDef, TColumn
