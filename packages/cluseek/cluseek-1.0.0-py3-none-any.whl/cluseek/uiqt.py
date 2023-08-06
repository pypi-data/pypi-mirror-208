import sys
import PySide2.QtWidgets as qtw
import PySide2.QtCore as qtc
import PySide2.QtGui as qtg
import math
import time
import webbrowser
import colorsys
import weakref
import enum
import re
import openpyxl as pxl
import csv
import os

import matplotlib
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure

import threading

from . import dbdl
from . import dframe
from . import about

DEBUG = False

# Import PySide2, the GUI library we're using

# DPI scaling adjustment, otherwise text will be too small on large screens
#   and too big on small screens -- without the rest of the UI scaling to match
if hasattr(qtc.Qt, 'AA_EnableHighDpiScaling'):
    qtw.QApplication.setAttribute(qtc.Qt.AA_EnableHighDpiScaling, True)

if hasattr(qtc.Qt, 'AA_UseHighDpiPixmaps'):
    # Not that there are any :)
    qtw.QApplication.setAttribute(qtc.Qt.AA_UseHighDpiPixmaps, True)

# GUI classes relying on templates
#   Please note that many parts of the UI are generated without
#   a template, or may not even have their own class.
from .uilayouts import resources
from .uilayouts.importframe import Ui_ImportFrame
from .uilayouts.filewidget import Ui_FileWidget
from .uilayouts.graphframe import Ui_GraphFrame
from .uilayouts.constraintsframe import Ui_ConstraintsFrame
from .uilayouts.accsetblock import Ui_AccsetBlock
from .uilayouts.neiviewer import Ui_NeiViewer
from .uilayouts.wrapperwindow import Ui_MainWindow
from .uilayouts.topclusters import Ui_TopClusters
from .uilayouts.proteininfo import Ui_ProteinInfo
from .uilayouts.clusterinfo import Ui_ClusterInfo
from .uilayouts.neiview_creator import Ui_NeiviewCreator
from .uilayouts.taxoninfo import Ui_TaxonInfo
from .uilayouts.regioninfo import Ui_RegionInfo
from .uilayouts.selector_widget import Ui_SelectorWidget
from .uilayouts.blastconfig import Ui_BlastConfig
from .uilayouts.dbprogressbar import Ui_DBProgressBar
from .uilayouts.excelexportframe import Ui_ExcelExportFrame

#Debug only
from collections import Counter
import code
import random
#code.interact(local=locals())

matplotlib.use("Qt5Agg")

class WorkerSignals(qtc.QObject):
    progress_manager_update = qtc.Signal(object, str, int)
    result = qtc.Signal(object)
    finished = qtc.Signal()
class WorkerThread(qtc.QThread):
    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()
        self.setTerminationEnabled(True)
        
        self.finished.connect(self.cbk_finished)
    @qtc.Slot()
    def run(self):
        result = self.func(*self.args, **self.kwargs)
        self.signals.result.emit(result)
        self.signals.finished.emit()
    def safe_start(self):
        assert AppRoot.work_thread is None, "Attempted to launch two simultaneous work threads"
        AppRoot.work_thread = self
        self.start()
    def cbk_finished(self):
        AppRoot.work_thread = None
        print("Work thread finished!")

class ProgressDialog(qtw.QDialog, Ui_DBProgressBar):
    class DotTimer(qtc.QObject):
        def __init__(self, master):
            super().__init__()
            self.timer_id = None
            self.master = master
        def start(self):
            self.stop()
            self.timer_id = self.startTimer(750)
        def stop(self):
            if self.timer_id:
                self.killTimer(self.timer.timer_id)
                self.timer_id = None
        def timerEvent(self, event):
            self.master.cbk_on_tick()
    def __init__(self, title, label_text, minimum=0, maximum=100, on_abort=None,
                 text_only=False):
        super().__init__()
        self.init_ui()
        self.reinitialize(title, label_text, minimum, maximum, on_abort,
                          text_only)
        self.timer = self.DotTimer(self)
        self.timer.start()
        self.trailing_dots = ""
    def init_ui(self):
        self.setupUi(self)
        self.abort_btn.clicked.connect(self.cbk_abort)
    def reinitialize(self, title, label_text, minimum=0, maximum=100, 
                     on_abort=None, text_only=False):
        # Static defaults
        self.canceled = False
        self.progress_pgbr.reset()
        self.setModal(True)
        
        self.trailing_dots = ""
        
        # Parameter-defined values
        self.setWindowTitle(title)
        self.setLabelText(label_text)
        self.progress_pgbr.setMaximum(maximum)
        self.progress_pgbr.setMinimum(minimum)
        self.on_abort = on_abort
        if text_only:
            self.abort_btn.setHidden(True)
            self.progress_pgbr.setHidden(True)
        else:
            self.abort_btn.setHidden(False)
            self.progress_pgbr.setHidden(False)
            
    def setLabelText(self, text):
        self.label_text = text
        self.info_lbl.setText(text)
    def cbk_abort(self):
        self.setLabelText("Aborting (Please be patient)")
        self.canceled = True
        self.on_abort()
    def wasCanceled(self):
        return(self.canceled)
    
    def setMaximum(self, val):
        self.progress_pgbr.setMaximum(val)
    def setMinimum(self, val):
        self.progress_pgbr.setMinimum(val)
    def setValue(self, val):
        self.progress_pgbr.setValue(val)
    def closeEvent(self, event):
        # We do not allow the users to close the window
        #   on their own. They have to sit and wait like
        #   good little users, so the app doesn't segfault.
        
        event.ignore()
        #self.show()
    def hideEvent(self, event):
        print("Hiding loading bar...")
    def on_abort(self):
        #To be overidden
        pass
    def cbk_on_tick(self):
        self.trailing_dots += "."
        if len(self.trailing_dots) > 3:
            self.trailing_dots = ""
        self.info_lbl.setText(self.label_text+" "+self.trailing_dots)
#other

def m_clear_layout(layout):
    #Adapted from qt docs
    #TODO: Can this spawn infinite recursion? God I hope not.
    while layout.count():
        child = layout.takeAt(0)
        if child.widget():
            child.widget().deleteLater()

#
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
 #                                                                   #
 #                         Abstract Classes                          #
 #                                                                   #
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 


class AppRoot():
    #This is effectively a container of global variables
    #  in the guise of a class definition.
    #  When you instantiate this class, the constructor modifies
    #  the base class (rather than the instance) to initialize all 
    #  the bells and whistles needed by other parts of the application.
    #
    #  AppRoot is called at will from anywhere within this module, to
    #  provide access to resources which there should never be more
    #  than one of, eg. database managers or Qt's QApplication class
    #  this is also why we use the base class instead of an instance
    #  (though an instance should work just as well for most purposes)
    #
    #  The reason we need to instantiate it first is simply because
    #  we don't want the module to start accessing the hard drive or
    #  building files the moment it is imported -- rather, it has to 
    #  be told to do so explicitly when we are ready for it.
    
    @classmethod
    def __init__(cls):
        #cls.qtapp
        #cls.ui_topclusters = None
        #cls.ui_importmanager = None
        #cls.ui_constraintsframe = None
        #cls.ui_neighborhoodviewer = None
        #cls.ui_infowindow = None
        
        #Colors. Yes really.
        cls.active_color = None
        #TODO: Is this actually being used though?
        #Timer
        cls.timer = TimerHandler()
        cls.work_thread = None
        cls.progress_dialog = None
    @classmethod
    def launch(cls):
        cls.qtapp = qtw.QApplication(sys.argv)
        cls.mainwindow = MainWindow()
        cls.mainwindow.show()
        cls.qtapp.exec_()
    @classmethod
    def on_app_start(cls):
        # Create progress dialog
        cls.progress_dialog = ProgressDialog(title="Launching...", 
                                             label_text="Please wait", 
                                             text_only=True)
        cls.progress_dialog.show()
        
        def on_db_init_complete():
            # * * Run other tasks * *
            # Making a decorator doesn't quite work.
            # It's not as explicit, but it works for now.
            UIProteinCluster._init_color_manager()
            ColorableButton._load_icons()
            cls.progress_dialog.setHidden(True)
        worker_db_init = WorkerThread(cls._on_app_start)
        worker_db_init.signals.finished.connect(on_db_init_complete)
        
        worker_db_init.safe_start()
        
    @classmethod
    def _on_app_start(cls): # * * Init databases * *
        cls.ipgdb = dbdl.ManagerIPG(
            "./data/ipgdb.csv", 
            overwrite=False, 
            verbose=True, 
            offline=False) #TODO:
        cls.neidb = dbdl.ManagerGB(
            "./data/neidb.txt.bz2",
            overwrite=False,
            verbose=True,
            offline=False)
        cls.cluster_manager = dbdl.ManagerClus()
    @classmethod
    def exit_application(cls):
        # Ew, but if the user wants to corrupt their data,
        #   c'est la vie. Better than the work thread staying
        #   alive in the background.
        if cls.work_thread:
            cls.work_thread.terminate()
        cls.qtapp.closeAllWindows()

class TimerHandler(qtc.QObject):
    # Q: Why does this exist?
    # A: Sometimes, we need to pass control over to
    #    Qt's loop for a bit to give it time to render --
    #    either because we're working on a time-intensive
    #    task, or because we need the geometry rendered 
    #    before modifying it further.
    def __init__(self):
        super().__init__()
        self.timers = {}
    def after_delay_do(self, callback, delay=10):
        # Start the timer and get its unique ID
        _id = self.startTimer(delay)
        
        assert _id != 0, "Timer failed to start for some reason."
        
        #Store the callback under the event's id
        self.timers[_id] = callback
    def timerEvent(self, event):
        # Save the event ID and callback function 
        #   in local variables
        _id = event.timerId()
        callback = self.timers[_id]
        
        # Clear the callback and stop the timer
        del self.timers[_id]
        self.killTimer(_id)
        
        # Execute the callback under the timer's ID
        # If this fails, the exception is raised, but we
        #   have already stopped the timer and cleared
        #   the callback, so the TimerHandler doesn't care.
        #   If we hadn't, the timer event would keep
        #   triggering over and over, spamming exceptions.
        callback()

class VariableTypes(enum.Enum):
    # Expect value
    EVALUE = "evalue"
    
    # Num identical aminoacids divided by length of hit sequence
    IDENTITY_PER_HIT = "identity_per_hit"
    
    # Num identical aminoacids divided by length of query sequence
    IDENTITY_PER_QUERY = "identity_per_query"
    
    # Length of alignment divided by length of query sequence
    #ALIGNMENT_PER_QUERY = "alignment_per_query"
    
    # Length of alignment divided by length of hit sequence
    #ALIGNMENT_PER_HIT = "alignment_per_hit"

class Dataset():
    class Rules():
        def __init__(self, dataset):
            self.dataset = dataset
            self.alignment_rules = {}
            self.sequence_rules = {}
            self.minimum_sequence_score = None
            for accset_name in self.dataset.accsets:
                self.alignment_rules[accset_name] = dict()
        def reset_sequence_rules(self):
            self.sequence_rules = {}
        def wipe_alignment_rule(self, accset_name, variable):
            if variable in self.alignment_rules[accset_name]:
                del self.alignment_rules[accset_name][variable]
        def new_alignment_rule(self, accset_name, variable, minimum=None, maximum=None, *args, **kwargs):
            rule = self.AlignmentRule(variable=variable, 
                                      minimum=minimum, 
                                      maximum=maximum)
            self.alignment_rules[accset_name][variable] = rule
        def get_alignment_rule(self, accset_name, variable):
            return(self.alignment_rules[accset_name].get(variable))
        def new_sequence_rule(self, accset_name, on_true):
            print(f"Rule {accset_name} with an on_true of {on_true} saved.")
            rule = self.SequenceRule(accset_name=accset_name, on_true=on_true)
            self.sequence_rules[accset_name] = rule
        def wipe_sequence_rule(self, accset_name):
            del self.sequence_rules[accset_name]
        def get_sequence_rule(self, accset_name):
            return(self.sequence_rules.get(accset_name))
        def set_minimum_sequence_score(self, min_score):
            self.minimum_sequence_score = min_score
        class Rule():
            # This is just a base class to identify rules of any kind.
            def __init__(self):
                pass
            def evaluate(self):
                pass
        class AlignmentRule(Rule):
            def __init__(self, variable, minimum, maximum, *args, **kwargs):
                super().__init__(*args, **kwargs)
                assert variable in VariableTypes, ("Invalid variable value. "
                                                    "Must be a VariableTypes enum.")
                self.variable = variable
                self.min = minimum # Inclusive
                self.max = maximum # Inclusive
            def evaluate(self, alignment):
                if self.variable is VariableTypes.EVALUE:
                    value = alignment.tophsp.expect
                elif self.variable is VariableTypes.IDENTITY_PER_QUERY:
                    value = alignment.tophsp.identities / alignment.record.query_length
                elif self.variable is VariableTypes.IDENTITY_PER_HIT:
                    value = alignment.tophsp.identities / alignment.length
                
                # We test the min max, and if neither fails, the rule succeeds.
                #   It shouldn't be possible for there to be neither min nor max,
                #   but technically it's not even such an issue.
                if self.min:
                    if value < self.min:
                        return(False)
                if self.max:
                    if value > self.max:
                        return(False)
                return(True)
        class SequenceRule(Rule):
            def __init__(self, accset_name, on_true):
                self.accset_name = accset_name
                self.on_true = on_true
    def __init__(self):
        #General
        self.names   = []
        self.accsets = {}
        self.blasts  = {}
        self.aliases = {}
        
        #Max bp between edges of features
        self.max_feature_distance = 60000
        #Max bp added to the edges
        self.border_size = 60000
        
        self.rules   = None
        self.root    = None
        self.subset  = None
        
        #The global alignment clustering result
        self.homo_clustering = None
        #The local alignment clustering result
        self.heth_clustering = None
    # * Primary Pipeline
    def load_files(self, paths):
        ftype = None
        xmls = []
        for path in paths:
            assert path.endswith(".xml")
            xmls.append(path)
        #Load XMLs
        records = dbdl.load_blastp_xmls(paths)
        for name in records:
            record = records[name]
            accset = set(record.als)
            
            self.names.append(name)
            self.accsets[name] = accset
            self.blasts[name] = record
            self.aliases[name] = record.query
        self.rules = self.Rules(self)
        self.build_root()
    def build_root(self):
        self.root = None
        allaccessions = []
        for accset in self.accsets.values():
            print(f"Added accset")
            allaccessions.extend(accset)
        self.root = AppRoot.ipgdb.build_ipg_tree(
            AppRoot.ipgdb.fetch_entries(allaccessions))
    # The following functions are part of a sequence
    #   sometimes I don't need to run the whole sequence,
    #   but regardless of where you stop, you should start
    #   from the top.
    def proc1_create_subsets(self):
        accsubsets = {}
        for accset_name in self.accsets:
            rules = [x for x in self.rules.alignment_rules[accset_name].values()]
            if not rules:
                accsubsets[accset_name] = set(self.blasts[accset_name].als)
            else:
                accsubsets[accset_name] = set()
            for alignment in self.blasts[accset_name].alignments:
                if all([rule.evaluate(alignment) for rule in rules]):
                    accsubsets[accset_name].add(alignment.accession)
        self.subset = {"accsets": accsubsets}
        print("Created accset subset using existing rules.")
    def proc2_extend_subset_using_ipg(self):
        not_found = 0
        added = 0
        for accsubset in self.subset["accsets"].values():
            to_add = set()
            for accession in accsubset:
                try:
                    to_add.update(self.root.ptAll[accession].ipt.pts)
                except KeyError:
                    not_found += 1
            accsubset.update(to_add)
            added += len(to_add)
        print("Extended accsubset using identical proteins."
             f"Added {added}, but {not_found} input accessions weren't"
              " found")
    def proc3_find_target_regions(self, region_filter=None):
        errors = Counter()
        
        # * We tag the marker proteins
        # First purge old marker tags:
        for pt in self.root.ptAll.values():
            if hasattr(pt, "marker"):
                pt.marker = set()
        # Then tag 'em anew:
        for accsubset_name in self.subset["accsets"]:
            accsubset = self.subset["accsets"][accsubset_name]
            for accession in accsubset:
                #TODO:STARTHERE: Account for sequence similarity between markers.
                #NOTE: Not sure if this is still relevant.
                try:
                    if hasattr(self.root.ptAll[accession], "marker"):
                        self.root.ptAll[accession].marker.add(accsubset_name)
                    else:
                        self.root.ptAll[accession].marker = {accsubset_name}
                except KeyError:
                    errors["Marker accession not found in root"] += 1
        
        hits = []
        
        
        for sc in self.root.scAll.values():
            errors["Sequences checked"] += 1
            # * Sort features by position
            features = sorted([ft for ft in sc.fts.values()],
                              reverse=False,
                              key=lambda ft: ft.start)
            
            # Keeps the score of a window. 
            #   If multiple windows merge, keeps best score from among them.
            window_best_score = 0
            last_window_best_score = 0
            window = []
            last_window = None
            hits_in_sc = []
            # We do a shifting window that is self.max_feature_distance large
            done = False
            while not done:
                # * Update window
                # add newest features
                errors["Windows checked"] += 1
                
                # As long as there's features left in the window,
                #   delete features from the left until we can add at least
                #   the next feature.
                while len(window) > 0 and ((features[0].stop - window[0].start) > self.max_feature_distance):
                    del window[0]
                
                added = 0
                # Fill the window up to capacity by adding features to the right
                while len(features)>0 and (len(window)==0 or ((features[0].stop - window[0].start) <= self.max_feature_distance)):
                    window.append(features[0])
                    del features[0]
                    added += 1
                
                # Flip the done flag at the end of the adding process
                #   if applicable
                if len(features) == 0:
                    done = True
                
                # The scoring variable
                passed = {}
                # Each key corresponds to one accset.
                #   The values are the scores assigned for that given accset.
                #   If multiple homologues from the same accset are present in
                #   a window, the score value is added only once.
                
                # Check rules
                for feature in window:
                    if hasattr(feature.ref, "marker"):
                        for accsubset_name in feature.ref.marker:
                            if accsubset_name in self.rules.sequence_rules:
                                passed[accsubset_name] = self.rules.sequence_rules[accsubset_name].on_true
                
                # If not all rules passed, just ignore this window and keep going.
                # OLD:
                #if not all(passed.values()): errors["Windows with no hits"] += 1; continue
                # NEW: The final score is the sum of all assigned scores.
                score = sum(passed.values())
                if score < self.rules.minimum_sequence_score:
                    # If score is BELOW minimum, we IGNORE it
                    #   (ergo, if score is == or greater, we USE it)
                    errors["Windows with no hits"] += 1
                    continue
                
                # Otherwise keep going.
                errors["Windows with hits"] += 1
                
                # * If windows in the same sequence overlap, merge them.
                # Check if this window overlaps with the last_window
                #   specifically, if this window starts before last_window stops
                
                #NOTE: I'm explicitly duplicating the int values in the
                #   following section, because some of them appear to have
                #   been getting kept, causing issues.
                if last_window and window[0].start < last_window[1]:
                    # Then extend the last_window with current one
                    last_window[1] = int(window[-1].stop)
                    errors["Extended window due to window hit overlap."] += 1
                    if score > last_window_best_score:
                        last_window_best_score = score
                else:# If there is no overlap,
                    #   add last_window to hits
                    if last_window:
                        hits_in_sc.append((int(last_window[0]), 
                                           int(last_window[1]), 
                                           int(last_window_best_score)))
                        errors["Non-overlapping hit in same sequence as an existing hit"] += 1
                    # and define current window as last_window
                    last_window_best_score = score
                    last_window = [int(window[0].start), int(window[-1].stop)]
            # Add the last window
            if last_window:
                hits_in_sc.append((int(last_window[0]), 
                                   int(last_window[1]), 
                                   int(last_window_best_score)))
            
            errors[f"Sequences with {len(hits_in_sc)} hit/s"] += 1
            # Generate target regions
            for start,stop,score in hits_in_sc:
                hits.append(UIGeneticRegion(self, sc, start, stop,
                                            hit_score=score))
            
            
        #TODO:
        #STARTHERE: see comments below. Need culling by taxon, and by run.
        
        if region_filter=="longest_per_taxon":
            # * Culling by taxon
            hits_by_taxon = {}
            # First we group hits together by run and taxon
            for hit in hits:
                # Get the taxon
                if hit.sc.tx not in hits_by_taxon:
                    hits_by_taxon[hit.sc.tx] = []
                hits_by_taxon[hit.sc.tx].append(hit)
            
            # Then we pick out only the longest region for each taxon
            passed_hits = []
            for tx in hits_by_taxon.values():
                # Get the hit that has the most features and only add that one
                #   to passed_hits
                longest = max(tx, key=lambda t: len(t.fts))
                errors[f"Hits culled due to multiple belonging to the same taxon."] += len(tx)-1
                passed_hits.append(longest)
            hits = passed_hits
        elif region_filter=="wgs_with_most_hits":
            # * Culling by run
            wgs_pattern = re.compile(r"(?P<wgs_run>[A-Z]{4,6}[0-9]{2})(?P<contig_id>[0-9]{6,99})")
            runs_by_taxon = {}
            # First we group hits together by run and taxon
            for hit in hits:
                # Get the taxon
                if hit.sc.tx not in runs_by_taxon:
                    runs_by_taxon[hit.sc.tx] = {}
                
                # Get the run name if it is a wgs sequence
                #   The big idea is that there's lesser chance of duplication when taking sequences
                #   from a single WGS run. We can imperfectly identify WGS runs based on their accessions. 
                # Accession formats: [number of letters]/[number of numerals]([total])
                #   Nucleotide - 1/5(6), 2/6(8), 2/8(10)
                #   Protein - 3/5(8), 3/7(10)
                #   WGS - 4 letters + 2 numerals for WGS assembly version followed by 6+ additional numerals
                #         6 letters + 2 numerals for WGS assembly version followed by 7+ numerals
                run_name = re.match(wgs_pattern, hit.sc.accession.split("_")[-1])
                if run_name is None:
                    run_name = hit.sc.accession
                if run_name not in runs_by_taxon[hit.sc.tx]:
                    runs_by_taxon[hit.sc.tx][run_name] = []
                runs_by_taxon[hit.sc.tx][run_name].append(hit)
            
            passed_hits = []
            for tx_runs in runs_by_taxon.values():
                # Get the run with most hits from each run and add it to hits, ditch everything else
                #   for that taxon. If several runs have the equal number of hits, keep the one
                #   which spans the most bp in total.
                
                # The run score is a float
                #   Above the decimal point, we evaluate the number of runs.
                #       This means the number of runs is superior to the second condition.
                #   Below the decimal point, we evaluate the combined length of all
                #       hits within a run.
                #       The expression 1/combined_length_of_runs will be smaller
                #       if the combined length is greater.
                #   If the number of runs is matched, the bp length will
                #       resolve the tie, rather than an arbitrary choice.
                best_run = max(list(tx_runs.values()), key=lambda run: len(run) - (1/sum([hit.length() for hit in run])))
                errors[f"While culling by run, added run with {len(best_run)} hits."] += 1
                passed_hits.extend(best_run)
            hits = passed_hits
        elif region_filter=="no_filter":
            pass
        else:
            raise ValueError("Invalid filter type specified in argument "
                            "region_filter. To use no filter, please use a "
                            "string reading 'no_filter'.")
            
        # * Finalize the results and save them.
        self.subset["hits"] = hits
        print("\nDone finding regions!")
        errors["Total hits found"] = len(hits)
        for key in errors:
            print(f"\t{key}: {errors[key]}")
        print("")
            
    def get_neighborhoods_for_subset(self):
        #First construct the request
        queries = set()
        for gr in self.subset["hits"]:
            #Request cds info for border_size bp around the cluster
            queries.add((gr.ref.sc.accession, 
                         max(gr.start - self.border_size, 1), 
                         gr.stop+self.border_size))
        # Make the request
        results = AppRoot.neidb.fetch_entries(queries)
        # Read the results and attach them to the contigs
        print(f"Adding features ...")
        t0 = time.time()
        times = []
        t1 = time.time()
        for result in results:
            AppRoot.neidb.add_features(result, self.root, types_=["cds"])
            t2 = time.time()
            times.append(t2 - t1)
            t1 = t2
        print(f"AVERAGE TIME: {sum(times)/max(len(times),1)} in {len(times)} records")
        print(f"Added features in {time.time() - t0} seconds!")

pass
#
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
 #                                                                   #
 #                         GUI Classes                               #
 #                                                                   #
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

class MainWindow(qtw.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
        # About info dialogs
        self.aboutme = None
        self.approot = AppRoot
        
        AppRoot.on_app_start()
    def init_ui(self):
        self.setupUi(self)
        
        self.setWindowTitle("CluSeek")
        
        # * License text popup
        self.licensetext = qtw.QTextBrowser()
        self.licensetext.setMinimumWidth(550)
        self.licensetext.setMinimumHeight(550)
        
        # * Menubar Callbacks
        self.action_aboutme.triggered.connect(self.cbk_aboutme)
        self.action_aboutqt.triggered.connect(self.cbk_aboutqt)
        
        self.action_offlinemode.triggered.connect(self.cbk_toggle_offlinemode)
        
        # Export actions
        self.action_export_filtergraph.triggered.connect(
            self.cbk_export_filtergraph)
        self.action_export_filterdata.triggered.connect(
            self.cbk_export_filterdata)
        self.action_export_neighborimage.triggered.connect(
            self.cbk_export_neighborimage)
        self.action_export_neighborexcel.triggered.connect(
            self.cbk_export_neighborexcel)
        
        # Neighborhood View config
        self.action_nei_proportional.triggered.connect(
            self.cbk_nei_proportional)
        self.action_nei_fixed.triggered.connect(
            self.cbk_nei_fixed)
        self.action_nei_setpropsizes.triggered.connect(
            self.cbk_nei_setpropsizes)
        self.action_nei_setfixedsizes.triggered.connect(
            self.cbk_nei_setfixedsizes)
        
        
        # * Other
        # This first
        AppRoot.ui_constraintsframe = ConstraintsFrame()
        self.lay_tab_filter.addWidget(AppRoot.ui_constraintsframe)
        
        # This second
        AppRoot.ui_importmanager = ImportManager()
        self.lay_tab_import.addWidget(AppRoot.ui_importmanager)
        
        # Finally these
        AppRoot.ui_neighborhoodviewer = NeighborhoodViewer()
        self.lay_tab_neighbors.addWidget(AppRoot.ui_neighborhoodviewer)
        AppRoot.ui_topclusters = TopClusters()
        self.lay_tab_test.addWidget(AppRoot.ui_topclusters)
        
        AppRoot.ui_infowindow = InfoWindow()
        
        
    def closeEvent(self, event):
        print("- - - Good night! - - -")
        AppRoot.exit_application()
    def cbk_aboutme(self):
        def callback_maker(textbox, text):
            def cbk_showlicense():
                textbox.setText(text)
                textbox.show()
            return(cbk_showlicense)
        if not self.aboutme:
            
            #scrollwrap = qtw.QWidget()
            scrollarea = qtw.QScrollArea()
            scrollable = qtw.QWidget()
            self.aboutme = scrollarea
            
            scrollarea.setWidget(scrollable)
            scrollarea.setMinimumWidth(900)
            scrollarea.setMinimumHeight(600)
            scrollarea.setWidgetResizable(True)
            scrollarea.setVerticalScrollBarPolicy(qtc.Qt.ScrollBarAlwaysOn)
            scrollarea.setHorizontalScrollBarPolicy(qtc.Qt.ScrollBarAlwaysOff)
            
            layout = qtw.QGridLayout()
            
            # About Me
            row=0
            label = qtw.QLabel(about.aboutme)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # About BLAST
            row+=1
            label = qtw.QLabel(about.aboutblast)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # About USEARCH
            row+=1
            label = qtw.QLabel(about.aboutusearch)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # About Python
            row+=1
            label = qtw.QLabel(about.aboutpython)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            #
            cbk_showlicense = callback_maker(self.licensetext, about.pylicense)
            btn = qtw.QPushButton("Python License")
            btn.clicked.connect(cbk_showlicense)
            layout.addWidget(btn, row, 4, 1, 1)
            
            # About PACKAGES
            row+=1
            label = qtw.QLabel(about.aboutpackages)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            label = qtw.QLabel(about.licensenote)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # PACKAGES
            row+=1
            headers = ["Module", "Version", "License Used*", "Copyright Notice/s", ""]
            for col in range(len(headers)):
                label = qtw.QLabel(headers[col])
                layout.addWidget(label, row, col)
            
            row+=1
            line = qtw.QFrame()
            line.setFrameShape(qtw.QFrame.HLine)
            line.setFrameShadow(qtw.QFrame.Sunken)
            layout.addWidget(line, row, 0, 1, 5)
            
            row+=1
            for package_info in about.get_aboutpackages():
                row += 1
                col =  0
                for value in package_info:
                    header = about.packagecsvheaders[col]
                    if header == "LicenseText":
                        cbk_showlicense = callback_maker(self.licensetext, value)
                        btn = qtw.QPushButton("License Text")
                        btn.clicked.connect(cbk_showlicense)
                        layout.addWidget(btn, row, col)
                    else:
                        layout.addWidget(qtw.QLabel(value), row, col)
                    col += 1
            
                        # PyThirdParty
            row+=1
            label = qtw.QLabel(about.aboutpythirdparty)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            headers = ["Module/Software", "skip", "skip", "Copyright Notice/s", ""]
            for col in range(len(headers)):
                if headers[col] == "skip":
                    continue
                label = qtw.QLabel(headers[col])
                layout.addWidget(label, row, col)
            
            row+=1
            line = qtw.QFrame()
            line.setFrameShape(qtw.QFrame.HLine)
            line.setFrameShadow(qtw.QFrame.Sunken)
            layout.addWidget(line, row, 0, 1, 5)
            
            row+=1
            for package_info in about.get_pythirdparty():
                row += 1
                col =  0
                for value in package_info:
                    header = about.packagecsvheaders[col]
                    if headers[col] == "skip":
                        col += 1
                        continue
                    if header == "LicenseText":
                        cbk_showlicense = callback_maker(self.licensetext, value)
                        btn = qtw.QPushButton("License Text")
                        btn.clicked.connect(cbk_showlicense)
                        layout.addWidget(btn, row, col)
                    else:
                        layout.addWidget(qtw.QLabel(value), row, col)
                    col += 1
            
            row+=1
            label = qtw.QLabel("\n\n")
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            
            
            # About PyInstaller
            row+=1
            label = qtw.QLabel("\n\n")
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            label = qtw.QLabel(about.aboutpyinstaller)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            label = qtw.QLabel(about.aboutpyinstaller2)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            label = qtw.QLabel(about.aboutpyinstaller3)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # PyInstaller Packages
            row+=1
            headers = ["Module", "Version", "License", "Copyright Notice/s", ""]
            for col in range(len(headers)):
                label = qtw.QLabel(headers[col])
                layout.addWidget(label, row, col)
            
            row+=1
            line = qtw.QFrame()
            line.setFrameShape(qtw.QFrame.HLine)
            line.setFrameShadow(qtw.QFrame.Sunken)
            layout.addWidget(line, row, 0, 1, 5)
            
            row+=1
            for package_info in about.get_aboutpyinstallerpackages():
                row += 1
                col =  0
                for value in package_info:
                    header = about.packagecsvheaders[col]
                    if header == "LicenseText":
                        cbk_showlicense = callback_maker(self.licensetext, value)
                        btn = qtw.QPushButton("License Text")
                        btn.clicked.connect(cbk_showlicense)
                        layout.addWidget(btn, row, col)
                    else:
                        layout.addWidget(qtw.QLabel(value), row, col)
                    col += 1
            
            scrollable.setLayout(layout)
        self.aboutme.setWindowTitle("About")
        self.aboutme.show()
    def cbk_aboutqt(self):
        meh = qtw.QWidget()
        aboutqt = qtw.QMessageBox.aboutQt(meh, "About Qt")
        meh.show()
    def cbk_toggle_offlinemode(self):
        if self.action_offlinemode.isChecked():
            AppRoot.ipgdb.set_offline(True)
            AppRoot.neidb.set_offline(True)
            print("Offline mode on!")
        else:
            AppRoot.ipgdb.set_offline(False)
            AppRoot.neidb.set_offline(False)
            print("Offline mode off!")
    # Export callbacks:
    def cbk_export_filtergraph(self):
        AppRoot.ui_constraintsframe.graphframe.export_graph_as_image()
    def cbk_export_filterdata(self):
        AppRoot.ui_constraintsframe.colocresultsframe.\
            export_as_spreadsheet()
    def cbk_export_neighborimage(self):
        AppRoot.ui_neighborhoodviewer.view.cbk_save_as_image()
    def cbk_export_neighborexcel(self):
        AppRoot.ui_neighborhoodviewer.view.export_into_excel()
    # Neiview callbacks
    def cbk_nei_proportional(self):
        AppRoot.ui_neighborhoodviewer.view.cbk_use_proportional_widths()
    def cbk_nei_fixed(self):
        AppRoot.ui_neighborhoodviewer.view.cbk_use_fixed_widths()
    def cbk_nei_setpropsizes(self):
        AppRoot.ui_neighborhoodviewer.view.configure_proportional_widths()
    def cbk_nei_setfixedsizes(self):
        AppRoot.ui_neighborhoodviewer.view.configure_fixed_widths()
    # Debug callbacks
    def cbk_codeinteract(self):
        code.interact(local=locals())

class ImportManager(qtw.QWidget, Ui_ImportFrame):
    # ----------- < < <   Subclass Definitions   > > > --------------
    class BlastConfig(qtw.QWidget, Ui_BlastConfig):
        # TODO: https://www.ncbi.nlm.nih.gov/books/NBK53758/
        #   This is a super useful page which defines the [categories]
        #   like [organism]
        default_settings = {
            "db": "nr",
            "entrez_query": "",
            "matrix_name": "BLOSUM62",
            "gapcosts": "11 1",
            "comp_based_statistics": 2,
            "nhits": 5000,
            "word_size": "6",
            "expect": 10**-10,
            "filter": "F"}
        def __init__(self):
            super().__init__()
            self.init_ui()
            
            # Copy defaults over settings.
            self.settings = dict(self.default_settings)
        def init_ui(self):
            self.setupUi(self)
            
            self.apply_btn.clicked.connect(
                self.cbk_apply)
            self.cancel_btn.clicked.connect(
                self.cbk_cancel)
            self.applydefaults_btn.clicked.connect(
                self.cbk_apply_defaults)
            
            # * Set up the initial values
            # First, input all the values into fields
        def cbk_apply(self):
            self.settings = self.read_values()
            self.close()
            print("Blast settings changed.")
            
        # Virtual function reimplementation
        def showEvent(self, event):
            self.sync_UI_and_saved_values()
            print("Blast settings shown.")
        
        def cbk_cancel(self):
            self.close()
        def sync_UI_and_saved_values(self):
            self.database_combo.setCurrentIndex(
                self.database_combo.findText(
                    "("+self.settings["db"]+")", 
                    qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
            
            self.entrez_ledit.setText("")
            
            self.matrix_combo.setCurrentIndex(
                self.matrix_combo.findText(
                    self.settings["matrix_name"], 
                    qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
            
            gapcosts = self.settings["gapcosts"].split(" ")
            gapcosts = f"Existence: {gapcosts[0]} Extension: {gapcosts[1]}"
            self.gapcost_combo.setCurrentIndex(
                self.gapcost_combo.findText(gapcosts, 
                    qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
            
            self.compadj_combo.setCurrentIndex(
                self.settings["comp_based_statistics"])
            
            self.maxhits_spin.setValue(self.settings["nhits"])
            
            self.wordsize_combo.setCurrentIndex(
                self.wordsize_combo.findText(
                    self.settings["word_size"], 
                    qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
            
            self.expect_spin.setValue(
                round(math.log(self.settings["expect"], 10)))
            
            filter = self.settings["filter"]
            self.mask_chk.setChecked("m" in filter)
            self.filter_chk.setChecked("T" in filter or "L" in filter)
        
        def cbk_apply_defaults(self):
            self.database_combo.setCurrentIndex(0)
            self.entrez_ledit.setText("")
            self.matrix_combo.setCurrentIndex(5)
            self.gapcost_combo.setCurrentIndex(2)
            self.compadj_combo.setCurrentIndex(2)
            self.maxhits_spin.setValue(5000)
            self.wordsize_combo.setCurrentIndex(2)
            self.expect_spin.setValue(-10)
            self.filter_chk.setChecked(False)
            self.mask_chk.setChecked(False)
            
        def read_values(self):
            values = {}
            
            # * Database
            values["db"] = self.database_combo.currentText().split("(")[-1]\
                .split(")")[0]
            
            # * Organisms
            values["entrez_query"] = self.entrez_ledit.text()
            
            # * Matrix
            values["matrix_name"] = self.matrix_combo.currentText()
            
            # * Gap Costs
            raw = " ".join(self.gapcost_combo.currentText().\
              lstrip("Existence: ").split(" Extension: "))
            values["gapcosts"] = raw
            
            # * Compositional adjustments
            values["comp_based_statistics"] = self.compadj_combo.currentIndex()
            
            # * Max target sequences
            values["nhits"] = self.maxhits_spin.value()
            
            # * Word size
            values["word_size"] = self.wordsize_combo.currentText()
            
            # * Expect threshold
            values["expect"] = 10 ** self.expect_spin.value()
            
            # ENSURE EXPECT VALUE IS A REAL NUMBER VIA ANOTHER CHECK
            
            
            # * Filter
            #    Low complexity filtering:
            #       F to disable, T or L to enable
            #    Masking:
            #       Prepend m for mask at lookup
            if self.filter_chk.isChecked():
                if self.mask_chk.isChecked():
                    values["filter"] = "mT"
                else:
                    values["filter"] = "T"
            else:
                values["filter"] = "F"
            
            return(values)

    class FileWidget(qtw.QWidget, Ui_FileWidget):
        #Represents a single file to be loaded
        FASTA1_pattern = re.compile(r"^>.+$")
        FASTA2_pattern = re.compile(r"^[A-Za-z]+$")
        accession_pattern = re.compile(r"^[A-Z]{3}[0-9]{5,7}(\.[0-9]+){0,1}$")
        
        def __init__(self, impman):
            super().__init__()
            self.impman = impman
            self.input_type = None # Need to setup UI first
            self.init_ui()
        def init_ui(self):
            self.setupUi(self)
            self.btn_browse.clicked.connect(self.add_files)
            self.btn_delete.clicked.connect(self.delete)
            
            # Adjust text field size based on text
            self.query_text.document().contentsChanged.connect(self.cbk_text_growth)
            
            # Change functionality when input type changes
            self.querytype_combo.currentIndexChanged.connect(
                self.cbk_input_type_combo_changed)
            
            # Check valid input
            self.query_text.textChanged.connect(
                self.cbk_text_edited)
                
            # Hide valid input error message until it is needed
            self.error_lbl.setHidden(True)
            
            self.setMinimumHeight(17+20)
            self.setMaximumHeight(17+20)
            self.updateGeometry()
            
            # First refresh to reflect the defaults
            self.cbk_input_type_combo_changed()
            
            # Set text window to resize up to a limit
        # * Handlers
        def update_minimum_height(self):
            height = (17
                +self.query_text.minimumHeight()
                +(0 if self.error_lbl.isHidden() else 30))
            self.setMinimumHeight(height)
            self.setMaximumHeight(height)
            self.updateGeometry()
        def cbk_text_growth(self):
            text_height = self.query_text.document().size().height()
            if text_height < 20:
                height = 20
            elif text_height < 115:
                height = text_height
            else:
                height = 115
            self.query_text.setMinimumHeight(height)
            self.query_text.setMaximumHeight(height)
            self.update_minimum_height()
        def cbk_input_type_combo_changed(self):
            value = self.querytype_combo.currentText()
            if value == "BLASTP XML":
                self.btn_browse.setEnabled(True)
                self.input_type = "XML"
            elif value == "NCBI protein accession":
                self.btn_browse.setEnabled(False)
                self.input_type = "Accession"
            elif value == "Amino acid FASTA":
                self.btn_browse.setEnabled(True)
                self.input_type = "FASTA"
            self.check_valid_input()
        def cbk_text_edited(self):
            self.check_valid_input()
        def add_file(self, path):
            # Set the path to the file as input, because this
            #   filetype is too big to display
            self.query_text.setText(path)
        def open_file(self, path):
            # Load the contents of the file as input
            with open(path, mode="r") as file:
                fasta = file.read()
            self.query_text.setText(fasta) 
            return
        def process_file(self, path):
            if self.input_type == "XML":
                self.add_file(path)
            elif self.input_type == "FASTA":
                self.open_file(path)
            return
        def add_files(self):
            if self.input_type == "XML":
                filter_string = "XML files (*.xml);;Any (*)"
            elif self.input_type == "FASTA":
                filter_string = "FASTA files (*.fsa *.fasta *.txt);;Any (*)"
            else:
                filter_string = "Error"
                print(f"Unknown input_type: {self.input_type}")
            
            paths = qtw.QFileDialog.getOpenFileNames(
                filter=filter_string)
            if not paths:
                # If the user cancelled out of the dialog, we peace out.
                return
            if isinstance(paths, tuple):
                paths = paths[0]
            else:
                # If the dialog is cancelled,
                #   just exit the function without changes.
                return
            
            # Load file / add path
            self.process_file(paths[0])
            
            # If multiple files are selected, open separate
            #   input boxes for each new one.
            if len(paths) > 1:
                for path in paths[1:]:
                    extrawidget = self.impman.add_filewidget()
                    extrawidget.querytype_combo.setCurrentIndex(self.querytype_combo.currentIndex())
                    extrawidget.process_file(path)
            return
        def delete(self):
            self.impman.filewidgets.remove(self)
            self.deleteLater()
            self.parent().adjustSize()
        def check_valid_input(self):
            text = self.query_text.document().toPlainText()
            
            # Don't scream at people until they actually type something
            if text == "": return
            
            # Red text html for screaming at people their inputs are wrong
            red_text_1 = '''<html><head/><body><p><span style=" color:#ff0004;">'''
            red_text_2 = '''</span></p></body></html>'''
            
            outcome = (True, None)
            if self.input_type == "XML":
                if not os.path.exists(text.replace("\n","")):
                    outcome = (False, red_text_1+"Invalid path"
                           +red_text_2)
            elif self.input_type == "Accession":
                if not re.match(self.accession_pattern, text):
                    outcome = (False, red_text_1+
                           "May not be a valid NCBI protein accession code"
                           +red_text_2)
            elif self.input_type == "FASTA":
                text = text.replace(" ","").replace("\t","")
                text = text.split("\n")
                if len(text) < 2:
                    failed = True
                elif not re.match(self.FASTA1_pattern, text[0]):
                    failed = True
                elif not re.match(self.FASTA2_pattern, "".join(text[1:])):
                    failed = True
                else:
                    failed = False
                if failed:
                    outcome = (False, red_text_1+"May not be valid FASTA"
                           +red_text_2)
            
            if not outcome[0]:
                self.error_lbl.setHidden(False)
                self.error_lbl.setText(outcome[1])
                self.update_minimum_height()
                return(outcome[0])
            else:
                self.error_lbl.setHidden(True)
                self.update_minimum_height()
                return(outcome[0])
        # in_filepath has been replaced with query_text
    # ----------- < < <   Method Definitions   > > > ----------------
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # * Processing vars
        self.filewidgets = [] #Stores the widgets storing path info
        
        self.dataset = None
        self.conframe = AppRoot.ui_constraintsframe
        self.blastconfig = self.BlastConfig()
        self.xmlpaths = []
        
        #Do this last
        self.init_ui()
    def init_ui(self):
        self.setupUi(self)
        #self.frame_neighborSettings.layout()\
        #    .setAlignment(qtc.Qt.AlignTop)
        self.lay_importFileStack.setAlignment(qtc.Qt.AlignTop)
        
        self.btn_addFile.clicked.connect(self.add_file)
        self.btn_loadFiles.clicked.connect(self.load_files)
        self.blastconfig_btn.clicked.connect(self.cbk_show_blastconfig)
        
        #This should happen last
        self.add_filewidget()        
        
        #Testing:
        #pal = qtg.QPalette(AppRoot.qtapp.palette())
        #pal.setColor(pal.Active, pal.Window, qtg.QColor(256, 96, 96))
        #self.setPalette(pal)
        
        #Testing2
        #self.setStyleSheet("background-color: cyan")
        
        #Testing3
        #Yes! Bingo! Very good. Dopamine kick.
        
        #Testing
        #self.btn_addFile.setStyleSheet(
        #    "QPushButton { "
        ##    "              border: 2px solid green;"
        #    "              border-color: green;"
        #    "              background-color: red;"
        ##    "              color: white;"
        #    "            }")
        #self.btn_addFile.setIcon(qtg.QIcon("./uilayouts/right_to_left.png"))
        #self.btn_loadFiles.setIcon(qtg.QIcon("./uilayouts/right_to_left.png"))
        
        
        #
        pass
    # * * Handlers * *
    def cbk_show_blastconfig(self):
        self.blastconfig.show()
    def load_files(self):
        # NOTE: The steps in this function are executed via work threads,
        #   and are declared in reverse order for this reason.
        
        AppRoot.progress_dialog.reinitialize(
          title="Neighborhood View Progress", 
          label_text="Waiting for Remote Blast (This may take a while)", 
          on_abort=None,
          text_only=True)
        
        # STEP-0: COLLECT INPUTS
        # * First we scrape the filewidgets for input data
        queries_to_blast = []
        self.xmlpaths = []
        print("Extracting information from filewidgets:")
        for filewidget in self.filewidgets:
            text = filewidget.query_text.document().toPlainText()
            print("\t"+filewidget.input_type+text)
            if text == "":
                continue
            elif filewidget.input_type == "XML":
                self.xmlpaths.append(text)
            elif filewidget.input_type == "FASTA"\
              or filewidget.input_type == "Accession":
                queries_to_blast.append(text)
        
        if len(self.xmlpaths) == 0 and len(queries_to_blast) == 0:
            print("Loading failed: No inputs provided!")
            return
        
        # STEP-3: DISPLAY DATA
        def on_files_loaded():
            #TODO: Temporary? Immediately load graph in main window.
            self.conframe.load_dataset(self.dataset)
            AppRoot.progress_dialog.setHidden(True)
        
        # STEP-2: LOAD DATA (And download IPG information)
        def on_remoteblast_complete():
            AppRoot.timer.after_delay_do(worker_loadfiles.safe_start)
        worker_loadfiles = WorkerThread(self._load_files)
        worker_loadfiles.signals.finished.connect(on_files_loaded)
        
        # With this somewhat confusing implementation, the
        #   download manager calls the signal in one thread,
        #   then responds to it in the other.
        worker_loadfiles.signals.progress_manager_update.connect(
            dbdl.DOWNLOAD_MANAGER.cbk_status_update)
        dbdl.DOWNLOAD_MANAGER.set_progress_manager_update_signal(
            worker_loadfiles.signals.progress_manager_update)
        
        # STEP-1: REMOTE BLAST
        worker_remoteblast = WorkerThread(self._run_remote_BLAST, 
                                          queries_to_blast)
        worker_remoteblast.signals.finished.connect(
            on_remoteblast_complete)
        

        # * Then we run the BLAST thread
        AppRoot.progress_dialog.show()
        worker_remoteblast.safe_start()
    def _run_remote_BLAST(self, queries_to_blast):
        if len(queries_to_blast) > 0:
            print("Sending remote BLAST request")
            # * Second, we need to perform remote BLASTs
            big_query = "\n".join(queries_to_blast)
            with open("./data/remote_blast_results.xml", mode="w") as file:
                file.write(
                    dbdl.BLASTManager.remote_blast(
                        big_query,
                        **self.blastconfig.settings).read())
            self.xmlpaths.append("./data/remote_blast_results.xml")
        else:
            print("Remote BLAST skipped, no queries.")
    def _load_files(self, progress_callback=None):
        self.dataset = Dataset()
        self.dataset.load_files(self.xmlpaths)
        print("Loaded files.")
    def add_file(self):
        self.add_filewidget()
    # * * Processors * *
    def add_filewidget(self):
        newfilewidget = self.FileWidget(self)
        self.lay_importFileStack.insertWidget(self.lay_importFileStack.count(), newfilewidget)
        self.filewidgets.append(newfilewidget)
        return(newfilewidget)
#[A, B]
class ConstraintsGraph(qtw.QWidget, Ui_GraphFrame):
    variable_texts = {
        "pquery": "% Identity",
        "evalue": "E-Value",
        "score": "Bit Score",
        }
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.xvals = []
        self.yvals = []
        
        self.dataset = None #Contains all 
        
        #Graph properties
        #self.graph = None #the graph itself
        
        self.accname = None #Active accset name
        self.acci = 0 #For cycling through accessions
        self.data = None #A tuple of values like evalue and %identity
        self.selectionbounds = None
        
        #Event handling (mouse events)
        self.pressedcoords = None
        
        #All possible variables
        self.variables = ["evalue", "pquery"]
        self.vari = 0 #For cycling through variables
        self.variable = self.variables[self.vari] #Active variable
        
        self.yscale = "linear" #Use log or linear
        
        
        self.init_ui()
    def switch_dataset(self, dataset):
        self.dataset = dataset
        self.change_accset(dataset.names[0])
    def init_ui(self):
        self.setupUi(self)
        
        # Initialize Graph
        self.fig = Figure(figsize=(5, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.graph  = FigureCanvasQTAgg(self.fig)
        
        self.le_min.editingFinished.connect(self.minmax_changed)
        self.le_max.editingFinished.connect(self.minmax_changed)
        
        self.lay_graphcontainer.addWidget(self.graph)
        
        # Hook up handlers
        self.btn_loglin.clicked.connect(self.cycle_scale)
        self.btn_variable.clicked.connect(self.cycle_variable)
        self.btn_addrule.clicked.connect(self.rule_make)
        self.btn_wiperule.clicked.connect(self.rule_wipe)
        
        #self.save_data_btn.clicked.connect(self.cbk_save_data)
        self.save_graph_btn.clicked.connect(self.cbk_save_graph)
        
        self.btn_variable.setText(self.variable)
        
        self.btn_loglin.setText("Linear")
        self.btn_variable.setText(self.variable_texts[self.variables[self.vari]])
    # * Handlers
    def minmax_changed(self):
        try:
            float(self.le_min.text())
        except ValueError:
            self.le_min.setText("")
        try:
            float(self.le_max.text())
        except ValueError:
            self.le_max.setText("")
        if self.le_min.text() != "" and self.le_max.text() != "":
            self.graph_setlimits(float(self.le_min.text()),
                                 float(self.le_max.text()))
    def change_accset(self, accname):
        self.accname = accname
        #Update the index so it cycles properly
        self.acci = self.dataset.names.index(accname)
        self.redraw()
    def cycle_variable(self):
        self.vari += 1
        if self.vari >= len(self.variables):
            self.vari = 0
        self.variable = self.variables[self.vari]
        variable_text = self.variable_texts[self.variable]
        self.btn_variable.setText(variable_text)
        self.redraw()
    def cycle_accset(self):
        self.acci += 1
        if self.acci == len(self.dataset.names):
            self.acci = 0
        self.change_accset(self.dataset.names[self.acci])
    def cycle_scale(self):
        if self.yscale == "linear":
            self.yscale = "log"
            self.btn_loglin.setText("Logarithmic")
        else:
            self.yscale = "linear"
            self.btn_loglin.setText("Linear")
        self.redraw()
    def graph_mousepress(self, event):
        if event.button == 1:
            # This only saves the position of the click 
            #    event for later
            self.pressedevent = event
    def graph_mouserelease(self, event):
        if event.button == 1:
            # The moment the button is released, this creates a 
            #   bounding box on the graph between where the mousepress
            #   and mouserelease happened.
            #print(f"MouseRelease-{event.button} at {event.xdata}, {event.ydata}")
            try:
                self.graph_setlimits(min(event.xdata, 
                                         self.pressedevent.xdata),
                                     max(event.xdata, 
                                         self.pressedevent.xdata))
            except TypeError:
                print("Drag and drop likely wasn't within bounds of the graph widget.")
        elif event.button == 3:
            self.graph_wipelimits()
    def graph_addhandlers(self):
        #self.graph.mousePressEvent = self.graph_mousepress
        #self.graph.mouseReleaseEvent = self.graph_mouserelease
        self.graph.mpl_connect("button_press_event", 
                               self.graph_mousepress)
        self.graph.mpl_connect("button_release_event", 
                               self.graph_mouserelease)
    def graph_wipelimits(self):
        if self.selectionbounds is not None:
            self.selectionbounds.remove()
            self.selectionbounds = None
            self.graph.draw()
        #TODO: Wipe current rule limits
        self.le_min.setText("")
        self.le_max.setText("")
        pass
    def graph_setlimits(self, xmin, xmax):
        if self.selectionbounds is not None:
            self.selectionbounds.remove()
            self.selectionbounds = None
        bounds = self.ax.get_xbound()
        self.selectionbounds = self.ax.axvspan(xmin, xmax, alpha=0.3, color="red")
        self.ax.set_xbound(bounds)
        self.graph.draw()
        #TODO: Set new rule limits
        self.le_min.setText(str(xmin))
        self.le_max.setText(str(xmax))
        pass
    def rule_make(self):
        xmin = float(self.le_min.text())
        xmax = float(self.le_max.text())
        rule = (xmin, xmax)
        #TODO: Try converting all the variable names to the new enum.
        #TODO: You know you want to.
        if self.variable == "evalue":
            variable = VariableTypes.EVALUE
        elif self.variable == "pquery":
            variable = VariableTypes.IDENTITY_PER_QUERY
        else:
            assert False, "Invalid ConstraintsGraph.variable value. Must be 'evalue' or 'pquery'"
        self.dataset.rules.new_alignment_rule(
            accset_name=self.accname, variable=variable, minimum=xmin, maximum=xmax)
        
        # After making the rule, recalc and redraw
        self.dataset.proc1_create_subsets()
        self.redraw()
    def rule_wipe(self):
        if self.variable == "evalue":
            variable = VariableTypes.EVALUE
        elif self.variable == "pquery":
            variable = VariableTypes.IDENTITY_PER_QUERY
        else:
            assert False, "Invalid ConstraintsGraph.variable value. Must be 'evalue' or 'pquery'"
        self.dataset.rules.wipe_alignment_rule(self.accname, variable)
        self.graph_wipelimits()
        
        # After making the rule, redraw
        self.dataset.proc1_create_subsets()
        self.redraw()
    def rules_load(self):
        #Updates the displayed limits 
        #  based on current accset and variable        
        if self.variable == "evalue":
            variable = VariableTypes.EVALUE
        elif self.variable == "pquery":
            variable = VariableTypes.IDENTITY_PER_QUERY
        else:
            assert False, "Invalid ConstraintsGraph.variable value. Must be 'evalue' or 'pquery'"
        rule = self.dataset.rules.get_alignment_rule(self.accname, variable)
        if rule is None:
            self.graph_wipelimits()
        else:
            self.graph_setlimits(rule.min, rule.max)
    def redraw_prepare_data(self, count_ipgs=True):
        # These are the three sets of values we'll be outputting.
        data_all = []
        data_subset = []
        data_in_hits = []
        name = self.accname
        errors = Counter()
        
        # Get all accessions in subset, if available
        try:
            subset = self.dataset.subset["accsets"][name]
        except:
            subset = set()
        
        # Get all accessions in hits, if available
        try:
            hit_accessions = set()
            hit_accessions_duplicated = list()
            for hit in self.dataset.subset["hits"]:
                hit_accessions.update(hit.fts)
                hit_accessions_duplicated.extend(list(hit.fts))
            hit_accessions = hit_accessions.intersection(self.dataset.subset["accsets"][name])
        except:
            hit_accessions = set()
            hit_accessions_duplicated = list()
        
        #for alignment in self.dataset.blasts[name].alignments:
        # This algorithm is done wrong. It needs to iterate over
        #   all HITS, and show everything on a per-hit basis.
        #   The reason is, accessions aren't unique, so counts
        #   get skewered relative to the totals otherwise.
        print(f"Sorting through {len(self.dataset.blasts[name].alignments)} alignments.")
        for alignment in self.dataset.blasts[name].alignments:
            hsp = max(alignment.hsps, key=lambda hsp: hsp.score)
            # Get the IPG
            if count_ipgs:
                try:
                    ipg_accs = set(self.dataset.root.ptAll[alignment.accession].ipt.pts)
                except KeyError:
                    errors["Accessions missing in data frame."] += 1
                    ipg_accs = {alignment.accession}
            else:
                ipg_accs = {alignment.accession}
            # This should calculate %query (pquery) as
            #   identities/query length
            # In order: Identities as % of query, evalue, bitscore,
            #   query_coverage, hit_coverage
            vals = [(hsp.identities / self.dataset.blasts[name].query_length,
                    hsp.expect,
                    hsp.score,
                    hsp.query_end-hsp.query_start+1,
                    hsp.sbjct_end-hsp.sbjct_start+1)]
            
            
            # All data
            data_all.extend(vals)
            # Subset after graph restrictions
            if ipg_accs.intersection(subset):
                data_subset.extend(vals)
            # Proteins found in actual potential clusters
            if ipg_accs.intersection(hit_accessions):
                data_in_hits.extend(vals)
        #code.interact(local=locals())
        print(f"Found {len(data_all)} alignments, {len(data_subset)} in subset and {len(data_in_hits)} in hits.")
        return(data_all, data_subset, data_in_hits)
    def redraw(self):
        #Redraws the graph in full
        
        zeroes = 0 # Keeps track of excluded data not shown on graph.
        zeroes_in_subset = 0
        zeroes_in_hits = 0
        
        data_all,data_subset,data_in_hits = self.redraw_prepare_data()
        
        if self.variable == "pquery":
            data_all     = [vals[0] for vals in data_all]
            data_subset  = [vals[0] for vals in data_subset]
            data_in_hits = [vals[0] for vals in data_in_hits]
            xbounds = {"lower": 0, "upper": 1}
            bins = [x/100 for x in range(0, 101)]
            xscale = "linear"
        elif self.variable == "evalue":
            data_all     = [vals[1] for vals in data_all]
            data_subset  = [vals[1] for vals in data_subset]
            data_in_hits = [vals[1] for vals in data_in_hits]
            # Drop all the zeroes. Sorry, no place for you in my graph, at least for now.
            #   Added note saying as much to graph title.
            for dataset in (data_all, data_subset, data_in_hits):
                for i in reversed(range(0,len(dataset))):
                    if dataset[i] == 0:
                        if dataset is data_all:
                            zeroes += 1
                        if dataset is data_subset:
                            zeroes_in_subset += 1
                        if dataset is data_in_hits:
                            zeroes_in_hits += 1
                        del dataset[i]
            
            
            # If all the hits are zeroes, we have a problem.
            try:
                bins = [10**x for x in range(math.floor(math.log(min(data_all), 10))-5, 
                                              math.ceil(math.log(max(data_all),10))+5)]
            
                xbounds = {"lower": min(bins+[1]), "upper": max(bins)}
                xscale="log"
            except ValueError:
                bins = 1
                xbounds = {"lower": 1, "upper": 2}
                xscale = "linear"
        elif self.variable == "score": # Unused
            data_all     = [vals[2] for vals in data_all]
            data_subset  = [vals[2] for vals in data_subset]
            data_in_hits = [vals[2] for vals in data_in_hits]
            try:
                xbounds = {"lower": min(data_all)-5, "upper": max(data_all)+5}
            except ValueError:
                xbounds = {"lower": None, "upper": None}
            bins = 1/0 #TODO: Fix this when you care.
            xscale = "linear"
            
        
        self.graph.deleteLater()
        
        self.fig = Figure(figsize=(5, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.graph = FigureCanvasQTAgg(self.fig)
        #TODO: Testing this
        self.graph_addhandlers() #For mouse picking on the graph
        self.rules_load() #To display previously saved rules
        
        self.lay_graphcontainer.addWidget(self.graph)
        
        total = len(data_all)+zeroes
        colocalized = len(data_in_hits)+zeroes_in_hits
        selected = len(data_subset)+zeroes_in_subset - colocalized
        ignored = total - selected - colocalized
        
        if ignored > 0:
            self.ax.hist(data_all, bins=bins, color="blue", alpha=0.5, 
                label=f"Unselected ({ignored})")
        self.ax.hist(data_subset, bins=bins, color="orange", alpha=1,
            label=f"Uncolocalized ({selected})")
        self.ax.hist(data_in_hits, bins=bins, color="red", alpha=1,
            label=f"Colocalized ({colocalized})")
        
        self.ax.legend()
        
        titletext = f"{self.dataset.aliases[self.accname]}"
        if zeroes:
            titletext += f"\nNote: evalues of zero (perfect match) not visible on graph (={zeroes})"
        
        
        self.ax.set_title(titletext)
        self.ax.set_ylabel("n proteins")
        self.ax.set_xlabel(self.variable_texts[self.variable])
        self.ax.set_xscale(xscale)
        self.ax.set_yscale(self.yscale)
        self.ax.set_xbound(**xbounds)
        
        self.ax.grid(True)
        self.graph.draw()
    def export_graph_as_image(self):
        path,suffix = qtw.QFileDialog.getSaveFileName(
                        caption="Save output",
                        filter="PNG(*.png);;JPEG(*.jpg *.jpeg)")
        if not path:
            return
        print(suffix)
        if suffix == "PNG(*.png)":
            self.graph.print_png(path)
        elif suffix == "JPEG(*.jpg *.jpeg)":
            self.graph.print_jpg(path)
    def cbk_save_graph(self):
        self.export_graph_as_image()
    def cbk_save_data(self):
        self.export_graph_data_as_spreadsheet()
    
class ColocResultsFrame(qtw.QTableWidget):
    def __init__(self, *args, **kwargs):
        super().__init__()
        self.coloc_table_basic = self
        self.table_initialized = False
        self.init_ui()
    def init_ui(self):
        pass
    def show_results(self, dataset):
        headers = ["Score", "Length (bp)", "Internal Length (bp)",
                   "Taxon", "Tax ID", "Sequence"]
        
        # Initialize table
        self.table_initialized = True
        self.coloc_table_basic.clear()
        
        self.coloc_table_basic.setRowCount(len(dataset.subset["hits"]))
        self.coloc_table_basic.setColumnCount(len(headers)+len(dataset.accsets))
        
        # Disable sorting so it doesn't interfere during generation
        self.coloc_table_basic.setSortingEnabled(False)
        
        # Headers
        accset_order = list(dataset.accsets)
        for accset_name in accset_order:
            headers.append(dataset.aliases[accset_name])
        self.coloc_table_basic.setHorizontalHeaderLabels(headers)
        
        # Values
        row = -1
        for hit in dataset.subset["hits"]:
            row += 1
            col = 0
            
            # Hitscore
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.hit_score)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Length
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.length())
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Internal Length
            fts = sorted([ft for ft in hit.fts.values()], key=lambda ft: ft.start)
            
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, fts[-1].start-fts[0].stop)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Taxon
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.sc.tx.sciname)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # TaxID
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.sc.tx.taxid)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Sequence
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.sc.accession)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Figure out which markers are present
            markers = {}
            for ft in hit.fts.values():
                if hasattr(ft.ref, "marker"):
                    for marker in ft.ref.marker:
                        if marker not in markers:
                            markers[marker] = []
                        markers[marker].append(ft.ref)
            for i in range(len(accset_order)):
                accset_name = accset_order[i]
                if accset_name in markers:
                    displaytext = ", ".join(
                        [pt.accession for pt in markers[accset_name]])
                else:
                    displaytext = ""
                    
                item = qtw.QTableWidgetItem()
                item.setData(qtc.Qt.DisplayRole, displaytext)
                item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
                self.coloc_table_basic.setItem(row, col+i, item)
            
        #Re-enable sorting
        self.coloc_table_basic.setSortingEnabled(True)
    def export_as_spreadsheet(self):
        if not self.table_initialized:
            print("Export failed: No spreadsheet initialized.")
            return
        
        path,suffix = qtw.QFileDialog.getSaveFileName(
                        caption="Save output",
                        filter="Excel 2010 Spreadsheet(*.xlsx)")
        
        if not path:
            return
        
        # self proxy
        t = self.coloc_table_basic
        
        # Initialize workbook
        wbk = pxl.Workbook()
        
        sh = wbk.active
        sh.title = "Colocalization Results"
        
        # copy headers
        for col in range(0, t.columnCount()):
            item = t.horizontalHeaderItem(col)
            sh.cell(column=col+1, row=1, value=item.text())
        
        # copy cells
        for row in range(0, t.rowCount()):
            for col in range(0, t.columnCount()):
                item = t.item(row,col)
                sh.cell(column=col+1, row=row+2, value=item.text())
        
        wbk.save(path)
        
class ConstraintsFrame(qtw.QWidget, Ui_ConstraintsFrame):
    # ----------- < < <   Subclass Definitions   > > > --------------
    class AccsetBlock(qtw.QWidget, Ui_AccsetBlock):
        def __init__(self, conframe, name):
            super().__init__()
            self.name = name
            
            #List of other AccsetBlocks
            self.conframe = conframe
            #  This widget will add itself to it, and then remove itself
            #  upon deletion, removes itself from the list
            self.conframe.accsetwidgets.append(self)
            
            self.alias = None
            
            #do this last
            self.init_ui()
        def init_ui(self):
            self.setupUi(self)
            self.ch_include.setText(self.name)
            
            self.btn_view.clicked.connect(self.selected)
            self.ch_include.stateChanged.connect(self.cbk_checked)
            
            # Set the default alias
            if self.name not in self.conframe.dataset.aliases:
                self.in_alias.setText(self.name)
            else:
                self.in_alias.setText(self.conframe.dataset.aliases[self.name])
            self.in_alias.textChanged.connect(self.cbk_renamed)
            self.in_alias.setCursorPosition(0)
        def easyscore_widgets_on(self):
            self.ch_include.setEnabled(True)
            self.score_spin.setEnabled(False)
            if self.score_spin.value() <= 0:
                self.ch_include.setChecked(qtc.Qt.Unchecked)
            else:
                self.score_spin.setValue(1)
        def easyscore_widgets_off(self):
            self.ch_include.setChecked(qtc.Qt.Checked)
            self.ch_include.setEnabled(False)
            self.score_spin.setEnabled(True)
        def selected(self):
            self.conframe.display_accset(self.name)
        def cbk_renamed(self):
            val = self.in_alias.text()
            if val == "":
                val = self.name
            self.conframe.dataset.aliases[self.name] = val
        def cbk_checked(self):
            if self.ch_include.checkState():
                if self.conframe.easyscore:
                    self.score_spin.setValue(1)
                    self.conframe.update_minscore()
            else:
                if self.conframe.easyscore:
                    self.score_spin.setValue(0)
                    self.conframe.update_minscore()
        def export_rules(self):
            # This should only be called when the user can
            #   no longer mess with the numbers.
            self.conframe.dataset.rules.new_sequence_rule(
                self.name,
                on_true=self.score_spin.value())
        def remove(self):
            self.conframe.accsetwidgets.remove(self)
            self.deleteLater()
            self.parent().adjustSize()
    class NeiviewCreator(qtw.QWidget, Ui_NeiviewCreator):
        def __init__(self, cframe):
            super().__init__()
            self.cframe = cframe
            self.init_ui()
        def init_ui(self):
            self.setupUi(self)
            
            # Link callbacks
            self.cancel_btn.clicked.connect(self.cbk_cancel)
            self.create_btn.clicked.connect(self.cbk_create)
            self.clus_local_evalue_input.editingFinished.connect(
                self.cbk_evalue_input_changed)
            self.chk_gray_out_rare.clicked.connect(
                self.cbk_gray_out_rare_changed)
                
            self.setWindowTitle("Creating new neighborhood view ...")
        def on_show(self):
            pass
        def cbk_create(self):
            # Borders
            self.cframe.dataset.border_size = \
                self.border_spin.value()
            
            # Settings
            settings = {
                "local_clustering_identity_threshold": round(self.clus_local_pidentity_spin.value()/100, 2),
                "local_clustering_evalue_threshold": float(self.clus_local_evalue_input.text()),
                "global_clustering_identity_threshold": round(self.clus_global_identity_spin.value()/100, 2),
                "flag_highlight_markers": True if self.chk_highlight_markers.checkState() is qtc.Qt.Checked else False,
                "flag_gray_out_rare": True if self.chk_gray_out_rare.checkState() is qtc.Qt.Checked else False,
                "gray_out_rare_threshold": (self.spin_gray_out_rare_threshold.value()/100),
                "flag_black_out_singletons": self.chk_black_out_singletons.isChecked(),
                "borders_size": self.border_spin.value(),
            }
            
            # Apply rules to ensure the data reflects the current
            #   state of the GUI.
            self.cframe.apply_rules()
            
            # Ping neighbourhood viewer
            AppRoot.ui_neighborhoodviewer.prepare_neighborhood(settings)
            
            #End the dialog
            self.hide()
        def cbk_cancel(self):
            self.hide()
        def cbk_gray_out_rare_changed(self):
            if self.cbk_gray_out_rare_changed.isChecked():
                self.spin_gray_out_rare_threshold.setDisabled(False)
            else:
                self.spin_gray_out_rare_threshold.setDisabled(True)
        def cbk_evalue_input_changed(self):
            try:
                float(self.clus_local_evalue_input.text())
            except ValueError:
                self.clus_local_evalue_input.setText("1.e-10")
        def refresh(self):
            # Display the window if hidden
            # and update values based on current dataset state
            
            #n neighbors
            self.nneighborhoods_displaylabel.setText(
                str(len(self.cframe.dataset.subset["hits"])))
            
            self.show()
    # ----------- < < <   Method Definitions   > > > ----------------
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.init_ui()
        
        self.dataset = None
        
        self.currentdisplay = None #None, self.graphframe, (, "group")
        self.graphframe = ConstraintsGraph()
        self.colocresultsframe = ColocResultsFrame()
        self.neiview_creator = self.NeiviewCreator(self)
        
        self.accsetwidgets = []
        
        self.intersections = None
    def init_ui(self):
        self.setupUi(self)
        
        self.lay_accsets.setAlignment(qtc.Qt.AlignTop)
        
        self.btn_applyrules.clicked.connect(self.cbk_search)
        
        self.spin_maxdist.valueChanged.connect(\
            self.update_max_feature_distance)
        self.minscore_spin.valueChanged.connect(\
            self.cbk_min_score_changed)
        
        self.easyscore_chk.stateChanged.connect(self.cbk_easyscore_toggle)
        
        self.btn_showneighbors.clicked.connect(self.cbk_show_neiview)
        
        self.btn_viewresults.clicked.connect(self.cbk_show_colocresults)
        
        # These two checkboxes simply aren't allowed to be checked at the same time.
        self.chk_only_one_wgs_run.clicked.connect(self.cbk_wgs_only_checked)
        self.chk_only_longest_per_taxon.clicked.connect(self.cbk_taxon_only_checked)
    def display_accset(self, accname):
        if accname in self.dataset.blasts:
            self.show_graph()
            self.graphframe.change_accset(accname)
        else:
            self.display_none()
    def update_colocresults(self):
        self.colocresultsframe.show_results(self.dataset)
    def show_colocresults(self):
        if self.currentdisplay is not self.colocresultsframe:
            if self.currentdisplay is not None:
                self.lay_contextdisplay.removeWidget(self.currentdisplay)
                self.currentdisplay.setHidden(True)
            self.lay_contextdisplay.addWidget(self.colocresultsframe)
            self.currentdisplay = self.colocresultsframe
            self.currentdisplay.setHidden(False)
    def show_graph(self):
        if self.currentdisplay is not self.graphframe:
            if self.currentdisplay is not None:
                self.lay_contextdisplay.removeWidget(self.currentdisplay)
                self.currentdisplay.setHidden(True)
            self.lay_contextdisplay.addWidget(self.graphframe)
            self.currentdisplay = self.graphframe
            self.currentdisplay.setHidden(False)
    def display_none(self):
        if self.currentdisplay is not None:
            self.lay_contextdisplay.removeWidget(self.currentdisplay)
            self.currentdisplay = None
    def update_max_feature_distance(self):
        #Updates the maximum spacing between selected features
        self.dataset.max_feature_distance = self.spin_maxdist.value()
    def load_dataset(self, dataset):
        print("Loading dataset...")
        self.dataset = dataset
        AppRoot.active_dataset = dataset
        
        self.graphframe.switch_dataset(self.dataset)
        #Update the maxspacing spinbox
        self.spin_maxdist.setValue(self.dataset.max_feature_distance)
        
        for block in list(self.accsetwidgets):
            #First remove all the old widgets
            #  We use list() to duplicate the list of accsetwidgets,
            #  as the original will change size during iteration
            #  due to things being removed from it.
            block.remove()
        for name in self.dataset.names:
            #Then add the new ones
            block = self.AccsetBlock(self, name)
            self.lay_accsets.addWidget(block)
            block.ch_include.setChecked(qtc.Qt.Checked)
        
        # Run the easyscore toggle callback once to switch to default state
        self.cbk_easyscore_toggle()
        AppRoot.mainwindow.maintabs.setCurrentIndex(1)
        
        # Disable the data view button
        self.btn_viewresults.setDisabled(True)
    def apply_rules(self):
        # Get Rules
        self.dataset.rules.reset_sequence_rules()
        for accwidget in self.accsetwidgets:
            accwidget.export_rules()
        # Define region_filter
        region_filter = None
        if self.chk_only_longest_per_taxon.checkState() is qtc.Qt.Checked:
            region_filter = "longest_per_taxon"
        elif self.chk_only_one_wgs_run.checkState() is qtc.Qt.Checked:
            region_filter = "wgs_with_most_hits"
        else:
            region_filter = "no_filter"
        
        print("Creating subset...", end="")
        self.dataset.proc1_create_subsets()
        self.dataset.proc2_extend_subset_using_ipg()
        self.dataset.proc3_find_target_regions(region_filter=region_filter)
        print(" Done!")
        
        print("Updating dataset stats...", end="")
        self.update_dataset_stats()
        print(" Done!")
        self.graphframe.redraw()
        self.update_colocresults()
    def update_dataset_stats(self):
        #proteins = []
        #for accset_name in self.dataset.accsets:
        #    proteins.append(f"\t{accset_name}: "
        #                    f"{len(self.dataset.accsets[accset_name])}\n")
        #proteins = "".join(proteins)
        info = (f"Found {len(self.dataset.subset['hits'])} regions "
                f"in {len(set([hit.sc.tx.taxid for hit in self.dataset.subset['hits']]))} NCBI Taxa")
        self.out_infolabel.setText(info)
    @property
    def easyscore(self):
        return(True if self.easyscore_chk.checkState() is qtc.Qt.Checked else False)
    def count_intersections(self):
        #To count combinations
        regions = Counter()
        errors = Counter()
        
        #So we can easily convert accessions to accsets
        accession_to_accset = {}
        for accset_name in self.dataset.accsets:
            for accession in self.dataset.accsets[accset_name]:
                accession_to_accset[accession] = accset_name
                
        #Now count the combinations
        region=[]
        for sc in self.dataset.root.scAll.values():
            errors["Sequences processed"]+=1
            last = None
            #Go through features from left to right
            for ft in sorted([feat for feat in sc.fts.values()], key=lambda x: x.start):
                if last:
                    if ft.start - last.stop <= self.dataset.max_feature_distance:
                        region.append(ft)
                    else:
                        #Use a tuple of alphabetically sorted accset names as
                        # the key for the region counter.
                        if len(region) > 1:
                            regions[tuple(sorted(\
                                [accession_to_accset[ft.ref.accession] for ft in region]))] += 1
                            #Refresh region
                        region = []
                last = ft
            #And also save region if the end of features is hit
            if len(region) > 1:
                regions[tuple(sorted(\
                    [accession_to_accset[ft.ref.accession] for ft in region]))] += 1
                #Refresh region
                region = []
        
        #Display
        print(errors)
        print(f"\nTotal regions size: {len(regions)}\n")
        #Then display the bad boys
        for key in sorted(list(regions), key=lambda x: len(x)):
            if regions[key] > 0:
                #Convert the key tuples to alias tuples
                aliases = []
                for accset_name in key:
                    aliases.append(self.dataset.aliases.get(accset_name))
                aliases = tuple(sorted(aliases))
                print(aliases, ":", regions[key], ",")
    def update_minscore(self):
        if not self.easyscore: return
        minscore = 0
        for accwidget in self.accsetwidgets:
            if accwidget.score_spin.value() > 0:
                minscore += 1
        self.minscore_spin.setValue(minscore)
    def cbk_show_colocresults(self):
        self.update_colocresults()
        self.show_colocresults()
    def cbk_min_score_changed(self):
        print("Maxscore changed.")
        self.dataset.rules.set_minimum_sequence_score(self.minscore_spin.value())
    def cbk_show_neiview(self):
        self.neiview_creator.refresh()
    def cbk_code_interact(self):
        code.interact(local=locals())
    def cbk_wgs_only_checked(self):
        if self.chk_only_longest_per_taxon.checkState() is qtc.Qt.Checked:
            self.chk_only_longest_per_taxon.setCheckState(qtc.Qt.Unchecked)
    def cbk_taxon_only_checked(self):
        if self.chk_only_one_wgs_run.checkState() is qtc.Qt.Checked:
            self.chk_only_one_wgs_run.setCheckState(qtc.Qt.Unchecked)
    def cbk_easyscore_toggle(self):
        if self.easyscore:
            self.minscore_spin.setEnabled(False)
            for accwidget in self.accsetwidgets:
                accwidget.easyscore_widgets_on()
        else:
            self.minscore_spin.setEnabled(True)
            for accwidget in self.accsetwidgets:
                accwidget.easyscore_widgets_off()
        self.update_minscore()
    def cbk_search(self):
        self.btn_viewresults.setDisabled(False)
        self.apply_rules()
    def contextMenuEvent(self, event):
        event.accept()
        menu = qtw.QMenu(self)

#TODO: Depricated
class DynamicContextMenu():
    # Menu elements are shared across a class.
    _menu_elements = []
    
    # To add menu elements, use which accept
    #   the argument "menu" as its first argument
    #   (obviously after implicit args)
    #   EG: def constructor(self, menu):
    #           action = menu.addAction("foo")
    #           action.triggered.connect(lambda: print("bar"))
    
    
    # Decorator for denoting menu elements.
    @classmethod
    def _element_constructor(cls, method):
        cls._menu_elements.append(method)
    
    # If for some accursed subclassing reason you want to remove
    #   a constructor retroactively, use this.
    @classmethod
    def _remove_element_constructor(cls, method):
        cls._menu_elements.remove(method)
    # Handler when a context menu signal is received
    def contextMenuEvent(self, event):
        #event.reason() .pos() .globalPos()
        
        print("contextMenuEvent")
        
        event.accept()
        menu = qtw.QMenu(self)
        
        #Run all the constructors
        for constructor in self._menu_elements():
            print("Running Constructor")
            constructor(menu)
        
        #Display the menu at event pos (most likely the 
        #   cursor position.)
        menu.popup(event.globalPos())
    
    #@DynamicContextMenu.element_constructor
    def __sample_constructor(self, menu):
        def my_callback_function():
            print("Callback triggered!")
        action = menu.addAction("MyActionDisplayName")
        action.triggered.connect(my_callback_function)

#
    # # # # # # # # # # # # # # # # # # #
    #    Class Overrides                #
    #        - UI functionality.        #
    #          for base dframe types.   #
    #        -> Genetic Region          #
    #        -> Protein Cluster         #
    # # # # # # # # # # # # # # # # # # #

class UIGeneticRegion(dbdl.dframe.GeneticRegion):
    # - - - < Constructor > - - -
    def __init__(self, dataset, *args, **kwargs):
        self.dataset = dataset
        
        #Decides whether to reverse the region's features when rendering
        self.flipped = False    
        #This is a weak set of live buttons and spacers
        self.cluster_layout = weakref.WeakSet()
        
        super().__init__(*args, **kwargs)
        
        self.borders = 0
    #@property
    #def fts_borders(self):
    #    fts = {}
    #    for ftname in self.sc.fts:
    #        ft = self.sc.fts[ftname]
    #        if ((ft.start <= max((self.start-self.dataset.border_size), 0)) and (ft.stop >= (self.stop+self.dataset.border_size))):
    #            fts[ftname] = ft
    #    return(fts)
    
    @property
    def fts_borders(self):
        expanded_borders = dframe.GeneticRegion(
            dataset=self.dataset, 
            scaffold=self.sc, 
            start=max(self.start-self.dataset.border_size,0), 
            stop=self.stop+self.dataset.border_size)
        return(expanded_borders.fts)
    
dbdl.dframe.GeneticRegion = UIGeneticRegion

class UIProteinCluster(dbdl.dframe.ProteinCluster):
    # - - - < Class Setup > - - -
    #Creating styles for buttons
    _selected_style_key = "none" #Holds the KEY for a given style,
                                 #the STYLE itself can be obtained in cls.styles
                                 #  The KEY is a tuple of Red, Green and Blue (0-255)
                                 #  it should be less processing intensive than strings
                                 #  when doing colour lookup or whatever.
                                 #=> Because color operations on hundreds of buttons
                                 #   are a performance bottleneck
    
    
    color_picker = None #Initialized later
    
    def _colorStyle(red, green, blue):
        #Change the text color based on the relative perceived luminosity 
        #   of the chosen color.
        #   Sadly, this won't work perfectly -- some older monitors
        #   will warp the colours so much this little trick won't really
        #   make the text contrast enough.
        if (0.2126*red + 0.72152*green + 0.0722*blue) >= 128:
            #Use white text
            text_color = "0,0,0"
        else:
            #Otherwise use black text
            text_color = "255,255,255"
        return(  "QPushButton {"
                f"    background-color: rgb({red}, {green}, {blue});"
                f"    color: rgb({text_color});"
                 "}")
    
    @classmethod
    def add_color_style(cls, red, green, blue):
        #finds/creates a corresponding style
        if (red, green, blue) in cls.styles:
            return
        cls.styles[(red, green, blue)] = cls._colorStyle(red, green, blue)
        
        
        
        #palette = qtg.QPalette(AppRoot.qtapp.palette())
        #brush = qtg.QBrush(qtg.QColor(red, green, blue, alpha))
        #brush.setStyle(qtc.Qt.SolidPattern)
        #palette.setBrush(qtg.QPalette.Active,  qtg.QPalette.Window, brush)
        #palette.setBrush(qtg.QPalette.Inactive, qtg.QPalette.Window, brush)
        #palette.setBrush(qtg.QPalette.Disabled, qtg.QPalette.Window, brush)
        #return(palette)
        pass
    @classmethod
    def _init_styles(cls):
        #This class method needs to be run after AppRoot.qtapp
        #   has been initialized, so we can get the app's styleSheet
        cp = cls._colorStyle
        cls.styles = {
            #Light                  #Full                  #Dark
            "lred":cp(255,170,170), "fred":cp(255,0,0),    "dred":cp(128,64,64),
            "lpur":cp(255,170,255), "fpur":cp(255,0,255),  "dpur":cp(128,64,128),
            "lblu":cp(170,170,255), "fblu":cp(0,0,255),    "dblu":cp(64,64,128),
            "lcya":cp(170,255,255), "fcya":cp(0,255,255),  "dcya":cp(64,128,128),
            "lgre":cp(170,255,170), "fgre":cp(0,255,0),    "dgre":cp(64,128,64),
            "lyel":cp(255,255,170), "fyel":cp(255,255,0),  "dyel":cp(128,128,64),
            "lgray":cp(170,170,170),"gray":cp(128,128,128),"dgray":cp(64,64,64),
            "white":cp(255,255,255),"black":cp(0,0,0),
            #"none": AppRoot.qtapp.styleSheet(),
        }
        cls.styles["none"] = cls.styles["white"]
        #Styles done
        cls.selected_style = "none"
    
    #This needs to be run once after the app itself is started.
    @classmethod
    def _init_color_manager(cls):
        #THIS IS RUN FROM AppRoot.on_app_start!
        
        #Create color picker
        cls.color_picker = qtw.QColorDialog()
        
        #Initialize styles
        cls._init_styles()
        
        #Flip init flag
        cls.first_init = False
    
    # - - - < Constructor > - - -
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # Buttons that belong to the cluster
        self.active_buttons = weakref.WeakSet()
        # We use weak references since the buttons tend to disappear a lot.
        #   When the buttons get deleted, if there are any references left
        #   that point to them, Qt hits a segfault.
        
        self.active_style = self.styles["white"]
        # TODO: Define a property that determines this dynamically
        self.human_name = None
        self.user_annotation = ""
        # TODO:
        
        # The user can merge and un-merge 
    
    # - - - < UI Handlers > - - -
        # this space for rent
    # - - - < Misc > - - -
    def get_name(self):
        if self.human_name:
            return(self.human_name)
        else:
            return(self._id)
    def set_name(self, newname): 
        self.human_name = newname
        self.update_button_names()
    def get_annotation(self):
        if self.user_annotation != "":
            return(self.user_annotation)
        else:
            for annotation,count in self._get_composition().most_common():
                if annotation not in {"hypothetical protein", "unknown"}:
                    return(annotation)
            return("unknown")
    # - - - < Button Management > - - -
    def add_colorable_button(self, btn):
        self.active_buttons.add(btn)
    #
    def update_button_colors(self):
        for button in self.active_buttons:
            button.update_color_style()
    def update_button_names(self):
        for button in self.active_buttons:
            button.update_displayed_text()
    #
    def toggle_cluster(self):
        #NOTE: self._selected_style_key is a class variable, not an instance var
        if self.active_style is self.styles[self._selected_style_key]:
            print("Coloured to None")
            #If the selected style is the same as this cluster's style,
            #uncolour it
            self.active_style = self.styles["none"]
        else:
            #Otherwise paint it the active colour
            self.active_style = self.styles[self._selected_style_key]
        self.update_button_colors()
        # TODO:
        
    @classmethod
    def color_picker_show(cls):
        #Shows a menu, lets the user input a color, and then sets it as selected,
        #  so that any clicked clusters are coloured with it
        color = cls.color_picker.getColor()
        if not color.isValid(): return False
        color = (color.red(), color.green(), color.blue())
        
        cls.add_color_style(*color)
        cls._selected_style_key = color
        return True
dbdl.dframe.ProteinCluster = UIProteinCluster


# This class is slotted into dbdl to control
#   download starting/progress.
#   Downloads are initiated by
#   the database managers, but must be approved
#   by the download manager.
#   This manager asks user for permission, and
#   reports download status.
class UIDownloadManager():
    def __init__(self):
        self.download_active = False
        self.size_of_download = None
        self.already_downloaded = None
        self.database = None
        self.progress_manager_update_signal = None
        self.abort = False
        self.abort_dialog = None
        self.abortlock = threading.Lock()
    def set_progress_manager_update_signal(self, signal):
        self.progress_manager_update_signal = signal
    
    def cbk_status_update(self, func, type, number_left):
        func(type, number_left)
    
    # Main thread side
    def ui_confirm_download_start(self):
        # OLD: To be removed
        #dialog = qtw.QMessageBox()
        #dialog.setText("Download starting!")
        #dialog.setInformativeText(
        #    f"Download from NCBI database?\n"
        #    f"There are {self.size_of_download} entries requested, "
        #    f"that are not available locally. If they are not downloaded,"
        #    f"the software will simply use whatever data is available locally.")
        #dialog.setStandardButtons(qtw.QMessageBox.Yes | qtw.QMessageBox.No)
        #response = dialog.exec_()
        #if response == qtw.QMessageBox.Yes:
        #    return True
        #else:
        #    return False
        
        # NEW: We do not ask for permission anymore to keep the
        #   pipeline going. The user can always interrupt.
        
        return True
    def _download_starting(self, type, number_to_download):
        print("Download Starting")
        # type = What kind of data is being downloaded. 
        #        Valid so far: "IPG", "GB" adding "BLASTP"
        # number_to_download = number of entries present locally,
        #                       which will need to be downloaded
        # False = Download is OK, proceed
        # True = Do not start download
        assert self.download_active==False, "Only one download can be running at the same time."
        self.download_active = True
        self.size_of_download = number_to_download
        self.already_downloaded = 0
        self.database = type
        self.abort = False
        self.abort_dialog = None
        
        # Ask the user whether he wants the download to happen.
        proceed = self.ui_confirm_download_start()
        abort = not proceed
        
        if not abort:
            print("Download starting...")
            # If this download is happening, create the progress dialog.
            AppRoot.progress_dialog.reinitialize(title = "Download Progress", 
                                              label_text="Downloading", 
                                              minimum=0, 
                                              maximum=self.size_of_download, 
                                              on_abort=self.on_abort)
            AppRoot.progress_dialog.show()
        elif abort:
            print("Download cancelled.")
            # If it's not happening, reset relevant variables.
            self.download_active = False
        return abort
    def on_abort(self):
        self.abortlock.acquire()
        self.abort = True
        self.abortlock.release()
    def _download_status(self, type, number_left_to_download):
        self.abortlock.acquire()
        if self.abort:
            print("Aborting...")
        else:
            self.already_downloaded = self.size_of_download - number_left_to_download
            AppRoot.progress_dialog.setValue(self.already_downloaded)
        # Same as before, True interrupts download.
        self.abortlock.release()
    def _download_ended(self, type, number_left):
        print("Download ended.")
        self.download_active = False
        AppRoot.progress_dialog.setValue(self.size_of_download)
        # It is the responsibility of the superior process
        #   to hide the progress_dialog.
        #AppRoot.progress_dialog.setVisible(False)
        self.progress_manager_update_signal = None
    # Work thread side
    #   These functions call their main thread counterparts
    #   in a thread-safe way.
    def download_starting(self, type, number_to_download):
        self.progress_manager_update_signal.emit(
            self._download_starting, type, number_to_download)
    def download_status(self, number_left_to_download):
        self.progress_manager_update_signal.emit(
            self._download_status, "", number_left_to_download)
        return(self.abort)
    def download_ended(self):
        self.progress_manager_update_signal.emit(
            self._download_ended, "", -1)
dbdl.DOWNLOAD_MANAGER = UIDownloadManager()

#    # # # # # # # # # # # # # # # # # # #
#    #           Info Classes            #
#    # # # # # # # # # # # # # # # # # # #

class RegionInfo(qtw.QWidget, Ui_RegionInfo):
    def __init__(self, region, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setupUi(self)
        
        self.seqacc_infolabel.setText(region.sc.accession)
        #self.seqlen_infolabel.setText(f"{len(region.scaffold)}")
        self.regionlen_infolabel.setText(f"{region.stop-region.start}")
        self.regionpos_infolabel.setText(f"{region.start}-{region.stop}")
        
        def open_genbank(self):
            webbrowser.open(f"https://www.ncbi.nlm.nih.gov/nuccore/{region.sc.accession}")
        self.genbank_btn.clicked.connect(open_genbank)

class TaxonInfo(qtw.QWidget, Ui_TaxonInfo):
    def __init__(self, taxon, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setupUi(self)
        
        self.taxid_infolabel.setText(taxon.taxid)
        
        self.sciname_infolabel.setText(taxon.sciname)
        
        self.strain_infolabel.setText(taxon.strain)
        
        def open_genbank(self):
            webbrowser.open(f"https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?id={taxon.taxid}")
        self.genbank_btn.clicked.connect(open_genbank)

class ProteinInfo(qtw.QWidget, Ui_ProteinInfo):
    def __init__(self, protein, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setupUi(self)
        
        # Fill in values.
        self.lbl_accession.setText(protein.accession)
        
        if protein.name: self.lbl_name.setText(protein.name)
        else: self.lbl_name.setText("N/A")
        
        if protein.type: self.lbl_annotation.setText(protein.type)
        else: self.lbl_annotation.setText("N/A")
        
        if protein.seq: self.txb_AAsequence.insertPlainText(protein.seq)
        else: self.txb_AAsequence.insertPlainText("N/A, no sequence found locally")
        
        def open_genbank(self):
            webbrowser.open(f"https://www.ncbi.nlm.nih.gov/protein/{protein.accession}")
        self.btn_genbank.clicked.connect(open_genbank)

class ClusterInfo(qtw.QWidget, Ui_ClusterInfo):
    def __init__(self, cluster, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setupUi(self)
        
        # Fill in values.
        self.lbl_clusterid.setText(cluster._id)
        self.ledit_rename.setText(cluster.human_name)
        self.lbl_nproteins.setText(str(len(cluster.members)))
        self.lbl_pidentity.setText(str(cluster.clustering.identity))
        self.txb_members.insertPlainText(", ".join(list(cluster.members)))
        
        self.ledit_userannotation.setText(cluster.user_annotation)
        
        count = cluster._get_composition()
        self.txb_annotations.insertPlainText("\n".join(
            [f"{key}: {count[key]}" for key in count]))
        
        def rename():
            cluster.set_name(self.ledit_rename.text())
        self.btn_rename.clicked.connect(rename)
        
        def change_annotation():
            cluster.user_annotation = self.ledit_userannotation.text()
        self.btn_applyuserannotation.clicked.connect(change_annotation)

class InfoWindow(qtw.QWidget):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.init_ui()
        self.contents = []
        self.pinned = qtc.Qt.Unchecked
        self.infowidget = None
    
    def display_info(self, feature=None, protein=None, cluster=None, region=None, taxon=None):
        # * Replace the infoview
        
        #First, dump the old one
        if self.infowidget:
            self.infowidget.setParent(None)
        
        #Then, create a new one
        self.infowidget = qtw.QWidget()
        
        #Make a layout for it
        ilayout = qtw.QVBoxLayout()
        ilayout.setAlignment(qtc.Qt.AlignLeft)
        ilayout.setContentsMargins(0,0,0,0)
        ilayout.setSpacing(8)
        
        # * Generate the contents!
        
        #Info Checkbox
        checkbox = qtw.QCheckBox("Window Pinned")
        checkbox.setCheckState(self.pinned)
        checkbox.stateChanged.connect(self.toggle_pin)
        ilayout.addWidget(checkbox)
        
        # Update the infowindow with whatever's been fed to it
        if feature:
            pass
            #No FeatureInfo yet
        if protein:
            ilayout.addWidget(ProteinInfo(protein=protein))
        if cluster:
            ilayout.addWidget(ClusterInfo(cluster=cluster))
        if region:
            ilayout.addWidget(RegionInfo(region=region))
        if taxon:
            ilayout.addWidget(TaxonInfo(taxon=taxon))
        
        # * Finally...
        
        # append a spacer so it doesn't get spread out when the user
        #   expands the window (or the contents contract)
        ilayout.addStretch()
        
        # Attach the layout to our infowidget
        self.infowidget.setLayout(ilayout)
        
        # And the infowidget to self
        self.layout().addWidget(self.infowidget)
        
        # Show the window (in case it had been hidden)
        self.show()
        # and put it on top
        self.activateWindow()
    
    def toggle_pin(self, state):
        # Toggle whether the window always stays on top or not
        if state:
            print("lmao checked")
            self.pinned = qtc.Qt.Checked
            #UNION with existing window flags
            self.setWindowFlags(
                self.windowFlags() | qtc.Qt.WindowStaysOnTopHint)
        else:
            print("lmao unchecked")
            self.pinned = qtc.Qt.Unchecked
            #AND NOT
            self.setWindowFlags(
                self.windowFlags() & ~qtc.Qt.WindowStaysOnTopHint)
        self.show()
    
    def init_ui(self):
        layout = qtw.QVBoxLayout()
        layout.setAlignment(qtc.Qt.AlignLeft)
        layout.setContentsMargins(0,0,0,0)
        layout.setSpacing(0)
        self.setLayout(layout)
        self.setWindowTitle("Info")

#    # # # # # # # # # # # # # # # # # # #
#    #   Neighborhood View and Related   #
#    # # # # # # # # # # # # # # # # # # #

# UI Remake Classes

class ColorableButton(qtw.QPushButton):
    _displaysize_compact = (35, 20) #default size
    #It might have been superceded by a dictionary
    #   in NeighbourhoodView
    #_displaysize_compact = (15,15)
    _got_icons = False
    
    #This needs to be run only once on app launch.
    @classmethod
    def _load_icons(cls):
        #THIS IS CALLED FROM AppRoot.on_app_start
        print("Strand Icons Loaded!")
        #Call this from AppRoot on app creation
        cls.icon_right_to_left = qtg.QIcon(":/icons/right_to_left.png")
        cls.icon_left_to_right = qtg.QIcon(":/icons/left_to_right.png")
    
    # * Initial Constructors *
    def __init__(self, homcluster=None, hetcluster=None,
                 feature=None, protein=None, uiregion=None,
                 *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        #TODO: Maybe split this class in two to avoid this argument combinations hell
        assert ((homcluster and hetcluster) or (uiregion and feature and protein)),(
            "Not enough arguments for ColorableButton."
            "Need to specify homcluster and hetcluster, or provide a"
            "uiregion, feature and protein to allow automatic lookup of the former.")
        self.hetcluster = hetcluster
        self.homcluster = homcluster
        self.protein    = protein
        self.feature    = feature
        
        self.uiregion = uiregion
        
        # If hetcluster and homcluster have not been set by default, go find them.
        if not homcluster:
            self.homcluster = self.uiregion.neighborview.neighborhood.dataset\
                      .homo_clustering.member_to_cluster.get(self.feature.ref.accession)
        else:
            self.homcluster = homcluster
        if not hetcluster:
            self.hetcluster = self.uiregion.neighborview.neighborhood.dataset\
                      .heth_clustering.member_to_cluster.get(self.feature.ref.accession)
        else:
            self.hetcluster = hetcluster
        
        if self.homcluster and self.hetcluster:
            #Register this button with its clusters
            self.hetcluster.add_colorable_button(self)
            self.homcluster.add_colorable_button(self)
        else:
            print(self.protein.accession, self.homcluster, self.hetcluster)
        
        self.active_style = None
        
        
        #Do this last.
        self._setup_ui()
    def _setup_ui(self):
        #Set the icon.
        if self.feature:
            #The icon denotes the direction/thread. If strand is flipped, the arrow should reverse as well.
            #  Technically, these conditionals should do the trick.
            if ((self.feature.strand == dframe.Strand.COMPLEMENTARY) ^ self.uiregion.genetic_region.flipped):
                #+ means right to left
                self.setIcon(self.icon_right_to_left)
            elif ((self.feature.strand == dframe.Strand.PRIMARY) ^ self.uiregion.genetic_region.flipped):
                #- means left to right
                self.setIcon(self.icon_left_to_right)
                #This is a hack to put the icon to the right side
                #    of the button, rather than the left by
                #    inverting the QPushButton's internal layout.
                #    If these buttons ever break, this is why.
                #    
                #    Big thanks to Mike's forum post on qtcentre.org
                self.setLayoutDirection(qtc.Qt.RightToLeft)
            else:
                print("Invalid strand/orientation.")
                1/0
        
        #Then we call the update functions.
        #These are separate from self._setup_ui(), because they need to be
        #   called even after the creation of this button.
        self.update_size()
        self.update_color_style()
        self.update_displayed_text()

    # * Other *
    def get_active_cluster(self):
        #TODO:
        #   If the homcluster/hetcluster is None, this CAN return None!
        #   Not sure if this is a bug of everything working as intended.
        #   It -should- be a bug -- each ColourableButton should have both defined.
        #   For now, update_displayed_text and update_color_style perform checks
        #       to mitigate this.
        if not self.uiregion:
            #TODO: This is a hack to allow the topcluster table to work.
            return(self.hetcluster)
        elif self.uiregion.neighborview.display_policy["cluster"] == "homcluster":
            return(self.homcluster)
        elif self.uiregion.neighborview.display_policy["cluster"] == "hetcluster":
            return(self.hetcluster)
        else:
            assert False, ("Invalid value for display_policy['cluster']!" 
                           "Can be 'homcluster' or 'hetcluster'")


    # * Update functions *
    def update_all(self):
        self.apply_size_policy()
        self.apply_color_style()
        self.apply_best_name()
    def update_size(self):
        # We decide the button's size based on
        #   - Where in the UI it's located
        #   - Current display_policy
        # Performing the check each time is somewhat inefficient, but
        #    it should be more flexible to work with.
        if self.feature and self.uiregion.neighborview.display_policy["size_mode"] == "proportional":
            #If the button has a feature, and size mode is proportional
            self.setFixedSize(
                self.uiregion.neighborview.bp_to_pixels(
                    self.feature.stop - self.feature.start),
                self._displaysize_compact[1])
        elif ((self.feature and 
          self.uiregion.neighborview.display_policy["size_mode"] == "fixed")):
            #If the size_mode is explicitly "fixed", OR there is no feature
            self.setFixedSize(*self.uiregion.neighborview.display_policy["size_fixed"])
        elif not self.feature and not self.uiregion:
            #There is nothing we can do.
            # TODO: It'd be great to be able to set fixed size without relying on
            # having a defined self.uiregion, but maybe it's better to just leave
            # it up to auto / topclusters
            pass
        else:
            assert True, "Invalid NeighborhoodView.display_policy['size_mode']!"
    def update_displayed_text(self):
        #The button inherits its name (displayed text) from a cluster
        #This uses the same logic as color style
        # Should get called on...
        # 1. Button creation (Done!)
        # 2. Cluster name change
        # 3. Cluster policy change
        active = self.get_active_cluster()
        if active:
            self.setText(self.get_active_cluster().get_name())
        else:
            self.setText("N/A")
    def update_color_style(self):
        #The button inherits its colour from a cluster.
        #This should get called...
        # 1. On button creation (Done!)
        # 2. On cluster color change
        # 3. On cluster policy change
        active = self.get_active_cluster()
        if active:
            if self.active_style is not active.active_style:
                self.active_style = active.active_style
                self.setStyleSheet(self.active_style)
        else:
            self.active_style = UIProteinCluster.styles["black"]
            self.setStyleSheet(self.active_style)

    # * Interface events / callbacks (cbks) *
    #   These are kept separate since they get called
    #   from all sorts of events. Unlike update functions,
    #   they reroute calls elsewhere.
    def cbk_show_info(self):
        AppRoot.ui_infowindow.display_info(
            cluster=self.get_active_cluster(),
            feature=self.feature,
            protein=self.protein)
    def cbk_highlight_homcluster(self): #TODO:
        self.homcluster.toggle_cluster()
    def cbk_highlight_hetcluster(self): #TODO:
        self.hetcluster.toggle_cluster()
    def cbk_highlight_active(self):
        self.get_active_cluster().toggle_cluster()
    def cbk_show_color_selector(self):
        # If colour chosen by color_picker is not valid 
        #   probably because (user cancelled out),
        #   do not change colours.
        if not UIProteinCluster.color_picker_show():
            return
        #Automatically color the dominant cluster
        self.cbk_highlight_active()
    def cbk_align_to_active(self):
        self.uiregion.neighborview.align_regions_to_cluster(self.get_active_cluster())
    def cbk_uiflip_region(self):
        self.uiregion.cbk_flip()
        
        AppRoot.timer.after_delay_do(self.uiregion.cbk_realign_to_last, delay=12)
    def cbk_select(self):
        AppRoot.ui_topclusters.select_cluster(self.get_active_cluster())
    def cbk_unselect(self):
        AppRoot.ui_topclusters.unselect_cluster(self.get_active_cluster())
    # * Interface events *
    def mousePressEvent(self, event):
        if event.button() == qtc.Qt.LeftButton:
            #For left mouse clicks (right mouse click is contextmenu)
            if event.modifiers() == qtc.Qt.CTRL:
                self.cbk_highlight_active()
            elif event.modifiers() == qtc.Qt.SHIFT:
                self.cbk_show_color_selector()
            elif event.modifiers() == qtc.Qt.ALT:
                self.cbk_uiflip_region()
            else:
                self.cbk_show_info()
        event.accept()
    def mouseDoubleClickEvent(self, event):
        event.accept()
        self.cbk_highlight_active()
    def contextMenuEvent(self, event):
        event.accept()
        menu = qtw.QMenu(self)
        
        #ACTION: Set Color
        action = menu.addAction("Choose color [Shift+LMB]")
        action.triggered.connect(self.cbk_show_color_selector)
        
        #ACTION: Highlight hetcluster
        action = menu.addAction("Use last color [Ctrl+LMB]")
        action.triggered.connect(self.cbk_highlight_hetcluster)
        
        #ACTION: Highlight homcluster
        #action = menu.addAction("Highlight Homcluster")
        #action.triggered.connect(self.cbk_highlight_homcluster)
        
        #ACTION: Align to active
        action = menu.addAction("Align to")
        action.triggered.connect(self.cbk_align_to_active)
        
        #ACTION: Show info
        action = menu.addAction("Info [LMB]")
        action.triggered.connect(self.cbk_show_info)
        
        #ACTION: Flip region
        action = menu.addAction("Flip region [Alt+LMB]")
        action.triggered.connect(self.cbk_uiflip_region)
        
        if self.get_active_cluster() not in AppRoot.ui_topclusters.selection:
            #ACTION: Select cluster
            action = menu.addAction("Select")
            action.triggered.connect(self.cbk_select)
        else:
            #ACTION: Unselect cluster
            action = menu.addAction("Unselect")
            action.triggered.connect(self.cbk_unselect)
        
        # Generate menu at click location
        menu.popup(event.globalPos())
class RegionUI():
    class CdsSpacer(qtw.QSpacerItem):
        def __init__(self, region, bp_gap_size):
            super().__init__(0,0)
            self.bp_gap_size = bp_gap_size
            self.region = region
        def go_proportional(self):
            self.changeSize(self.region.neighborview.bp_to_pixels(self.bp_gap_size),0)
        def go_fixed(self):
            self.changeSize(0,0)
    # Container for the head and tail widgets which get integrated into the UI.
    #   The RegionUI can regenerate/modify these widgets.
    def __init__(self, genetic_region, neighborview):
        self.genetic_region = genetic_region
        self.neighborview = neighborview
        self.spacer = None
        
        # This a set of weak references to buttons within this
        #    instance. The list is generated and remade in self.make_tail_layout.
        self.buttons = weakref.WeakSet()
        # We use weak references since the buttons tend to disappear a lot.
        #   When the buttons get deleted, if there are any references left
        #   that point to them, Qt hits a segfault.
        
        self.head = None
        self.tail = None
        self.tail_spacers = weakref.WeakSet()
        
        self._setup_ui()
    def _setup_ui(self):
        #Prepare head and tail
        self.tail = qtw.QWidget()
        self.remake_tail()
        
        self.head = qtw.QWidget()
        self.head.setLayout(self.make_head_layout())
    def make_head_layout(self):
        head_layout = qtw.QHBoxLayout()
        head_layout.setAlignment(qtc.Qt.AlignRight)
        head_layout.setContentsMargins(0,0,0,0)
        head_layout.setSpacing(0)
        
        #Fill it out
        head_layout.addWidget(self.get_contig_button())
        head_layout.addWidget(self.get_taxon_button())
        return(head_layout)
    def make_tail_layout(self):
        # * * Prepare Tail layout
        # * Get all the proteins to display in the tail.
        features_to_display = []
        self.tail_spacers = weakref.WeakSet()
        for ft in self.genetic_region.fts_borders.values():
            if ft.type == "cds":
                features_to_display.append(ft)
        features_to_display.sort(key=lambda ft: ft.start, 
                                 reverse=self.genetic_region.flipped)
        # Prepare the tail layout
        tail_layout = qtw.QHBoxLayout()
        tail_layout.setAlignment(qtc.Qt.AlignLeft)
        tail_layout.setContentsMargins(0,0,0,0)
        tail_layout.setSpacing(0)
        
        #Add alignment spacer
        self.spacer = qtw.QWidget()
        tail_layout.addWidget(self.spacer)
        self.spacer.setFixedSize(1,1)
        
        # Create cluster buttons
        last = None
        
        for ft in features_to_display:
            # * Inserting gaps
            
            # ignore if this is the first cds
            if last:
                
                # need to account for tails being generated flipped
                if self.genetic_region.flipped:
                    spacer = self.CdsSpacer(self, last.start-ft.stop)
                else:
                    spacer = self.CdsSpacer(self, ft.start-last.stop)
            
                if self.neighborview.display_policy["size_mode"] == "proportional":
                    # If proportional, order spacer to switch to proportional
                    spacer.go_proportional()
                elif self.neighborview.display_policy["size_mode"] == "fixed":
                    # The spacer defaults to 0 width, do nothing.
                    pass
            
                # Remember the tail spacer
                self.tail_spacers.add(spacer)
            
            #Create a button widget for the cds feature
            btn = ColorableButton(
                feature=ft,
                protein=ft.ref,
                uiregion=self)
            tail_layout.addWidget(btn)
            self.buttons.add(btn)
            last = ft
        return(tail_layout)
    #
    def get_contig_button(self):
        btn = qtw.QPushButton(text = self.genetic_region.sc.accession)
        
        #Match height with cluster buttons
        #TODO: Make prettier, allow for changes in size
        btn.setFixedHeight(self.neighborview.display_policy["size_fixed"][1])
        
        btn.clicked.connect(self.cbk_contigbtn_clicked)
        return(btn)
    def get_taxon_button(self):
        #Pick which name to display. If no sciname is present, show taxid.
        if self.genetic_region.sc.tx.sciname:
            text = self.genetic_region.sc.tx.sciname
        else:
            text = self.genetic_region.sc.tx.taxid
        
        btn = qtw.QPushButton(text = text)
        
        #Match height with cluster buttons
        #TODO: Make more elegant, allow for changes in size
        btn.setFixedHeight(self.neighborview.display_policy["size_fixed"][1])
        
        btn.clicked.connect(self.cbk_txbtn_clicked)
        return(btn)
    #
    def cbk_contigbtn_clicked(self):
        AppRoot.ui_infowindow.display_info(region=self.genetic_region, 
                                           taxon=self.genetic_region.sc.tx)
    def cbk_txbtn_clicked(self):
        AppRoot.ui_infowindow.display_info(taxon=self.genetic_region.sc.tx)
    def cbk_flip(self):
        # Just flip the flag and leave the rest to existing functions.
        self.genetic_region.flipped = not self.genetic_region.flipped
        self.remake_tail()
    def remake_tail(self):
        AppRoot.timer.after_delay_do(self._remake_tail, delay=0)
    def _remake_tail(self):
        # Delete the old layout by reparenting it to a widget
        #   that is immediately deleted.
        
        if self.tail.layout():
            # This check lets us run this on init 
            # when the layout doesn't exist yet
            
            # First we need to clear the layout to avoid a segfault
            #m_clear_layout(self.tail.layout())
            
            # Then we reparent the layout,
            #   and hopefully no memory leak happens.
            qtw.QWidget().setLayout(self.tail.layout())
        
        # Replace it with the new layout
        self.tail.setLayout(self.make_tail_layout())
        # WARNING: There CAN NOT be any strong references left to
        #   *any* of the member buttons, otherwise you're going to get a segfault.
        #   This may mean this function cannot be called through a button-based callback.
        #   If there are any strong references, THE APPLICATION WILL SEGFAULT.
        
        # Switch tabs to neiviewer. This is necessary, because otherwise
        #   the autoalign of tails on creation won't work.
        #   TODO: This may be a temporary solution.
        AppRoot.mainwindow.maintabs.setCurrentIndex(2)
        # Then we call the aligner, which relies on actual pixel positions to work
        #AppRoot.timer.after_delay_do(self._align_to_centre, delay=0)
    #Utility
    def set_offset(self, offset):
        # Used for alignments
        self.spacer.setFixedSize(offset,1)
    def get_offset_to_centre(self):
        # Returns the offset to centre the sequence
        central_btn = sorted(list(self.buttons), key=lambda btn: btn.feature.start)[round(len(self.buttons)/2)]
        return(self.get_alignment_offset(central_btn.get_active_cluster()))
    def get_alignment_offset(self, cluster):
        #TODO: Accounting for multiples of the same cluster
        hit = None
        #Figure out the position of the button we want to align to
        for button in self.buttons:
            try:
                if ((button.hetcluster is cluster) or (button.homcluster is cluster)):
                    hit = button
                    break
            except:
                pass
        #print(clusterids)
        
        
        if not hit:
            #If we failed to find the cluster, offset is None
            desired_offset = 0
            return(None)
        else:
            # Otherwise, calculate offset:
            # * New method (pixel-based)
            # centre_offset defines the offset that needs to be applied for
            #   this region to align at width=0
            # minimum_offset is the minimum offset necessary for the spacer
            #   to have a positive width (IE so that buttons left of our
            #   aligned cluster don't end up outside the window)
            #centre_offset = -hit.geometry().left() + self.spacer.geometry().width()
            minimum_offset = hit.geometry().left() - self.spacer.geometry().width()
        
        return(minimum_offset)
    def get_length(self, ft):
        if self.neighborview.display_policy["size"] == "proportional":
            return(self.neighborview.bp_to_pixels(ft.stop-ft.start))
    def cbk_realign_to_last(self):
        # Used if the UI of this region is edited, causing it to
        #   likely go out of alignment.
        print("Realigning to last.")
        self.neighborview.align_to_last()
class NeighborhoodView(qtw.QWidget, Ui_NeiViewer):
    display_policy = {
        "size_mode": "proportional", #Can be "proportional" or "fixed"
        "size_fixed": (50,20), #The height-width to use when size_mode is "fixed"
        "cluster": "hetcluster", #Can be "homcluster" or "hetcluster"
        "ignore_small": 0, #Do not display clusters with fewer than n members
        "bp_per_pixel": 33} #This is the conversion factor between bp and pixels.
        #                    adds up to approximately 30 pixels per 1000 bp
    

    def __init__(self, neighborhood, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        self.neighborhood = neighborhood
        
        self.data_loaded = False
        
        # Genetic Regions that comprise the displayed subset
        self.displayed_regions = None
        
        # RegionUIs visible on screen right now
        self.active_regionuis = []
        
        # Cluster statistics for currently displayed neighborhoods
        self.cluster_instances_unique = None
        self.cluster_instances_total = None
        
        self.settings = {}
        
        self.excel_exporter = None
        
        self.init_ui()
        
        
        # Holds the last 5 clusters that were aligned to.
        # The latest one is always at index [-1]
        self.last_aligned_to = []
        self.last_offset = None
    def init_ui(self):
        self.setupUi(self)

        self.btns_added = 0
        self.dummies = [] #debug
        
        # * Make sure the contents of the layouts stack to top
        self.lay_clustercontainer.setAlignment(qtc.Qt.AlignTop)
        self.lay_headercontainer.setAlignment(qtc.Qt.AlignTop)
        
        # * Replace the horizontal scrollbar with a custom one
        self.botscroll = self.clustercontainer_scrl.horizontalScrollBar()
        self.updownscroll = self.vertiscroll_scrl.verticalScrollBar()
        # then reattach it elsewhere
        self.clustercontainer.adjustSize()
        self.lay_hscrollbarcontainer.addWidget(self.botscroll)
        
        #Make the internal scrolltable ignore scroll events
        #  so it can't get out of alignment.
        def ignore_event(event):
            event.ignore()
        self.clustercontainer_scrl.wheelEvent = ignore_event
        
        #Then attach a visibility updater
        #self.updownscroll.sliderMoved.connect(self.update_visibility)
        #tt = time.time()
        #print("Making dummies")
        #for i in range(5000):
        #    self.add_dummy(i)
        #print(f"Added {self.btns_added} dummies in {time.time()-tt}s.")
        pass
    #This utility method converts bp to pixels.
    def bp_to_pixels(self, bp):
        return(round(bp/self.display_policy["bp_per_pixel"]))
    def display_neighborhoods(self, settings=None, regions_to_display=None):
        # Determine if we're using a supplied subset of regions,
        #   or the output of search.
        if not regions_to_display:
            regions_to_display = self.neighborhood.dataset.subset["hits"]
        
        self.displayed_regions = regions_to_display
        
        # Compute the indidence of protein clusters
        self.calculate_cluster_counts()
        
        if settings:
            self.settings.update(settings)
        if self.settings.get("flag_gray_out_rare"):
            self.gray_out_rare(self.settings.get("gray_out_rare_threshold"))
        if self.settings.get("flag_black_out_singletons"):
            self.black_out_singletons()
        if self.settings.get("flag_highlight_markers"):
            self.highlight_marker_genes()
        self.separate_identically_named_clusters()
        
        # * Wipe old neighbourhood view:
        for i in reversed(range(self.lay_clustercontainer.count())):
            if self.lay_clustercontainer.itemAt(i).widget():
                self.lay_clustercontainer.itemAt(i).widget().setParent(None)
        for i in reversed(range(self.lay_headercontainer.count())): 
            if self.lay_headercontainer.itemAt(i).widget():
                self.lay_headercontainer.itemAt(i).widget().setParent(None)
        for i in range(20):
            time.sleep(0.25)
            print(".", end=" ")
        print("Done.")
        
        t0 = time.time()
        
        # * Pre-cook the contents of all the regions
        rc = {} #rc stands for region contents
        accessions_to_clusters = self.get_active_clustering().member_to_cluster
        for region in self.displayed_regions:
            contents = set()
            for feature in region.fts_borders.values():
                if feature.type == "cds":
                    contents.add(accessions_to_clusters.get(feature.ref.accession))
            rc[region] = contents
        
        # * Actually create the UI
        #   starter region
        region = self.displayed_regions[0] #ERROR HAPPENS HERE #STARTHERE:
        #   its sorted features, for deciding orientation
        region_sclusters = [accessions_to_clusters.get(ft.ref.accession)\
                            for ft in sorted(region.fts_borders.values(), key=lambda _ft: _ft.start)]
        regions_to_render = set(self.displayed_regions)
        regions_in_display_order = []
        
        while len(regions_to_render) > 0:
            
            regions_in_display_order.append(region)
            regions_to_render.remove(region)
            
            # Draw region
            regionui = RegionUI(genetic_region=region, neighborview=self)
            # NOTE: active_reguionuis is also being used as a reference when displaying
            #   the results in excel.
            self.active_regionuis.append(regionui)
            self.lay_clustercontainer.addWidget(regionui.tail)
            self.lay_headercontainer.addWidget(regionui.head)
            
            #
            
            # Decide next region
            rcA = rc[region]
            highest_region_score = -1
            next_region = None
            for regionB in regions_to_render:
                rcB = rc[regionB]
                score = len(rcB & rcA) / len(rcB | (rcA))
                #fwd_score = len(rcB & rcA) / (len(rcB & rcA) + len(rcB - rcA))
                #rev_score = len(rcB & rcA) / (len(rcB & rcA) + len(rcA - rcB))
                if score > highest_region_score:
                    highest_region_score = score
                    next_region = regionB
            
            # I can't wrap my head around this loop, so adding a break here
            #   for when we run out of regions. Sorry to whoever has to
            #   deal with this later (probably me -- sorry, future me).
            if not next_region:
                break
            
            
            # Orientation Determination
            #   We'll be using a flip index to do this one.
            #   Basically, we look at all the orientation of all the common
            #   clusters across both regions, and check if each pair is pointing
            #   in the same direction, or opposite directions between the two
            #   regions.
            #   EG:             Clus1   Clus2   Clus3   Clus4
            #           RegA     ->     -> <-    ->      <-
            #           RegB     <-     -> <-    <-      ->
            #           Subscore -1      0      -1      1
            #           Total
            common_clusters = rcA.intersection(rc[next_region])
            
            cluster_orientation = {}
            for cluster in common_clusters:
                #[0] - Strand sub-index in new_region
                #[1] - Total members in new_region
                #[2] - Strand sub-index in old_region
                #[3] - Total members in old_region
                cluster_orientation[cluster] = [0,0,0,0]
            
            for ft in sorted(next_region.fts_borders.values(), key=lambda _ft: _ft.start):
                cluster = accessions_to_clusters.get(ft.ref.accession)
                if cluster in common_clusters:
                    if ft.strand is dframe.Strand.PRIMARY:
                        cluster_orientation[cluster][0] += 1
                        cluster_orientation[cluster][1] += 1
                    elif ft.strand is dframe.Strand.COMPLEMENTARY:
                        cluster_orientation[cluster][0] -= 1
                        cluster_orientation[cluster][1] += 1
                    else:
                        assert False, f"Invalid strand value in a CDS feature ({ft.strand})."
            for ft in sorted(region.fts_borders.values(), key=lambda _ft: _ft.start):
                cluster = accessions_to_clusters.get(ft.ref.accession)
                if cluster in common_clusters:
                    if ft.strand is dframe.Strand.PRIMARY:
                        cluster_orientation[cluster][2] += 1
                        cluster_orientation[cluster][3] += 1
                    elif ft.strand is dframe.Strand.COMPLEMENTARY:
                        cluster_orientation[cluster][2] -= 1
                        cluster_orientation[cluster][3] += 1
                    else:
                        assert False, f"Invalid strand value in a CDS feature ({ft.strand})."
            
            flip_index = 0
            for cluster in cluster_orientation:
                tab = cluster_orientation[cluster]
                #Calculate the index component and add it up
                if (tab[0] > 0) == (tab[2] > 0):
                    # If the orientation of cluster instances in region and 
                    #   next_region are on average in the same direction,
                    
                    # Increment the index with the smaller of the total cluster counts
                    if hasattr(cluster, "_temp_is_marker"):
                        # If this is a marker (as labelled by the highlight markers method)
                        #   weight the flip score.
                        flip_index += min(tab[1], tab[3])*4
                    else:
                        flip_index += min(tab[1], tab[3])
                else:
                    # Otherwise, decrement it by the same
                    if hasattr(cluster, "_temp_is_marker"):
                        # And again, if this is a marker
                        #   weight the flip score.
                        flip_index -= min(tab[1], tab[3])*4
                    else:
                        flip_index -= min(tab[1], tab[3])
            
            if flip_index < 0:
                # If the majority of shared clusters are pointing 
                #   in opposite directions, flip the orientation.
                next_region.flipped = True if not region.flipped else False
            else:
                # Otherwise, leave it be.
                next_region.flipped = False if not region.flipped else True
            
            # End Loop
            region = next_region
        print(
            len(set(self.displayed_regions)), 
            len(set(self.displayed_regions).intersection(regions_in_display_order)))
        self.displayed_regions = regions_in_display_order
        
        t1 = time.time()
        print(f"Calculated region order and orientation in {t1-t0} seconds.")
        self.data_loaded = True
    def highlight_marker_genes(self):
        accessions_to_clusters = self.get_active_clustering().member_to_cluster
        
        colors = []
        for prefix in ["l", "d", "f"]:
            for colorname in ["red", "blu", "pur", "cya", "gre", "yel"]:
                colors.append(UIProteinCluster.styles[prefix+colorname])
        colori = 0
        
        marked_overall = set()
        for accset_name in self.neighborhood.dataset.accsets:
            accset = self.neighborhood.dataset.accsets[accset_name]
            
            #Load alias if available.
            if self.neighborhood.dataset.aliases[accset_name] != "":
                alias = self.neighborhood.dataset.aliases[accset_name]
            else:
                alias = accset_name
            
            #Cycle through the list of colors.
            color = colors[colori]
            colori += 1
            if colori == len(colors):
                colori = 0
            
            #Check which clusters the marker proteins ended up in...
            marked_in_accset = set()
            for accession in accset:
                marked_in_accset.add(accessions_to_clusters.get(accession))
            marked_in_accset.remove(None)
            
            #Then rename all those clusters.
            for cluster in marked_in_accset.intersection(marked_overall):
                #If a cluster has two or more markers, combine the names
                cluster.human_name = " + ".join([cluster.human_name, accset_name])
                #And set color. I don't plan to do separate colors for hybrids.
                cluster.active_style = color
                
                # Just a label for autoflipping, for the love of god don't use
                #   this for anything else!!!
                cluster._temp_is_marker = None
            for cluster in marked_in_accset.difference(marked_overall):
                #If a cluster has only one marker, simply set the name to alias
                cluster.human_name = alias
                #And set color
                cluster.active_style = color
                
                # Just a label for autoflipping, for the love of god don't use
                #   this for anything else!!!
                cluster._temp_is_marker = None
                
            
            #Then update the list of already marked.
            marked_overall = marked_overall.union(marked_in_accset)
    def gray_out_rare(self, threshold):
        #TODO: Add a check so this doesn't override manually coloured clusters!
        #NOTE: Done, I think. In the ugliest way I could think of.
        threshold = math.floor(len(self.displayed_regions)*threshold)
        for cluster in self.neighborhood.dataset.homo_clustering.clusters.values():
            if self.cluster_instances_unique[cluster] <= threshold:
                if any((
                  cluster.active_style is UIProteinCluster.styles["none"],
                  cluster.active_style is UIProteinCluster.styles["gray"],
                  cluster.active_style is UIProteinCluster.styles["dgray"])):
                    cluster.active_style = UIProteinCluster.styles["gray"]
        for cluster in self.neighborhood.dataset.heth_clustering.clusters.values():
            if self.cluster_instances_unique[cluster] <= threshold:
                if any((
                  cluster.active_style is UIProteinCluster.styles["none"],
                  cluster.active_style is UIProteinCluster.styles["gray"],
                  cluster.active_style is UIProteinCluster.styles["dgray"])):
                    cluster.active_style = UIProteinCluster.styles["gray"]
    def black_out_singletons(self):
        for cluster in self.neighborhood.dataset.homo_clustering.clusters.values():
            if self.cluster_instances_unique[cluster] == 1:
                if any((
                  cluster.active_style is UIProteinCluster.styles["none"],
                  cluster.active_style is UIProteinCluster.styles["gray"],
                  cluster.active_style is UIProteinCluster.styles["dgray"])):
                    cluster.active_style = UIProteinCluster.styles["dgray"]
        for cluster in self.neighborhood.dataset.heth_clustering.clusters.values():
            if self.cluster_instances_unique[cluster] == 1:
                if any((
                  cluster.active_style is UIProteinCluster.styles["none"],
                  cluster.active_style is UIProteinCluster.styles["gray"],
                  cluster.active_style is UIProteinCluster.styles["dgray"])):
                    cluster.active_style = UIProteinCluster.styles["dgray"]
    def calculate_cluster_counts(self):
        clustering = self.get_active_clustering()
        accessions_to_clusters = clustering.member_to_cluster
        
        # * Count unique and total counts.
        #A count of regions with at least one instance
        cluster_instances_unique = Counter()
        #A dict of counters for adding up the total number
        #   of instances, as well as the number of duplications
        cluster_instances_total = {} 
        #TODO: Debugging check, maybe remove later.
        failed_to_find_cluster = 0
        
        for cluster in clustering.clusters.values():
            #Set up the cluster instances counter
            cluster_instances_total[cluster] = Counter()
        
        
        for region in self.displayed_regions:
            unique_clusters_in_region = set()
            total_clusters_in_region = Counter()
            for feature in region.fts_borders.values():
                if feature.type != "cds": continue
                
                try:
                    cluster = accessions_to_clusters[feature.ref.accession]
                except KeyError:
                    failed_to_find_cluster+=1
                    continue
                
                unique_clusters_in_region.add(cluster)
                total_clusters_in_region[cluster] += 1
            
            # Increment the counter for unique instances
            for cluster in unique_clusters_in_region:
                cluster_instances_unique[cluster] += 1
            # Add up the number of instances for non-unique cases
            for cluster in total_clusters_in_region:
                #For example, if 8 regions have cluster occur 2 times, then
                #   cluster_instances_total[cluster][2] will equal 8
                cluster_instances_total[cluster][total_clusters_in_region[cluster]] += 1
        
        print(f"Failed to identify the clusters of {failed_to_find_cluster} "
               "features while creating overview table.")
        
        self.cluster_instances_unique = cluster_instances_unique
        self.cluster_instances_total = cluster_instances_total
    def separate_identically_named_clusters(self):
        human_names = {}
        for cluster in self.get_active_clustering().clusters.values():
            # If the cluster has been renamed
            if cluster.human_name:
                name = cluster.human_name
                if cluster.human_name in human_names:
                    # but if there already has been a cluster with
                    #   the same name, append an index like (1), (2), ... 
                    #   to differentiate them
                    cluster.set_name(f"{name} ({human_names[name]})")
                    human_names[name] += 1
                else:
                    human_names[name] = 1
    def change_display_policy(self):
        pass
        #TODO:
    def update_tails(self):
        for regionui in self.active_regionuis:
            regionui.remake_tail()
    def align_to_last(self):
        if self.last_aligned_to:
            self.align_regions_to_cluster(self.last_aligned_to[-1])
    def cbk_use_proportional_widths(self):
        # This can get called either when we're switching from
        #   fixed to proportional, OR
        #   when we're changing widths
        if self.display_policy["size_mode"] != "proportional":
            self.display_policy["size_mode"] = "proportional"
        
        for region in self.active_regionuis:
            # Update spacer widths
            for spacer in region.tail_spacers:
                spacer.go_proportional()
            # Update button widths
            for button in region.buttons:
                button.update_size()
        
        # Align all to last
        AppRoot.timer.after_delay_do(self.align_to_last, 250)
        self.cbk_on_proportions_changed()
    def cbk_use_fixed_widths(self):
        # This can get called either when we're switching from
        #   proportional to fixed, OR
        #   when we're changing widths
        # Set fixed
        if self.display_policy["size_mode"] != "fixed":
            self.display_policy["size_mode"] = "fixed"
            # If switching from something that's not fixed,
            #   update both gaps and buttons
            for region in self.active_regionuis:
                for spacer in region.tail_spacers:
                    spacer.go_proportional()
                for button in region.buttons:
                    button.update_size()
        else:
            # If we're just resizing button width,
            #   only update buttons.
            for region in self.active_regionuis:
                for button in region.buttons:
                    button.update_size()
        
        # Align all to last
        AppRoot.timer.after_delay_do(self.align_to_last,250)
        self.cbk_on_proportions_changed()
    def configure_proportional_widths(self):
        #STARTHERE:
        # Get fixed width from user.
        bpratio = qtw.QInputDialog.getInt(
            self, "Base pairs per pixel",
            "Set how many base pairs a single pixel represents:",
            value=self.display_policy["bp_per_pixel"],
            minValue=1)
            
        # QInputDialog.getInt() returns a tuple (int, bool)
        #   The int is the value we want, the bool denotes
        #   whether the dialog got accepted, or cancelled. 
        #   We only proceed if the user didn't cancel out.
        if not bpratio[1]: return
        
        self.display_policy["bp_per_pixel"] = bpratio[0]
        
        # Update the view
        if self.display_policy["size_mode"] == "proportional":
            self.cbk_use_proportional_widths()
    def configure_fixed_widths(self):
        # Get fixed width from user.
        widthvalue = qtw.QInputDialog.getInt(
            self, "Fixed With",
            "Input the display width of proteins (tiles), in pixels.",
            value=self.display_policy["size_fixed"][0],
            minValue=1)
            
        # QInputDialog.getInt() returns a tuple (int, bool)
        #   The int is the value we want, the bool denotes
        #   whether the dialog got accepted, or cancelled. 
        #   We only proceed if the user didn't cancel out.
        if not widthvalue[1]: return
        
        oldsize = self.display_policy["size_fixed"]
        newsize = (widthvalue[0], oldsize[1])
        self.display_policy["size_fixed"] = newsize
        
        # Update the view
        if self.display_policy["size_mode"] == "fixed":
            self.cbk_use_fixed_widths()
    def cbk_on_proportions_changed(self):
        state = self.display_policy["size_mode"] == "proportional"
        AppRoot.mainwindow.action_nei_proportional.setChecked(state)
        AppRoot.mainwindow.action_nei_fixed.setChecked(not state)
    def align_regions_to_cluster(self, cluster):
        # * Collect offsets
        offsets = {}
        for regionui in self.active_regionuis:
            offsets[regionui] = regionui.get_alignment_offset(cluster)
        #
        self.minimum_offset = max([offset if offset is not None else 0 for offset in offsets.values()])
        
        for regionui in offsets:
            offset = offsets[regionui]
            if offset is None:
                regionui.set_offset(0)
            else:
                regionui.set_offset(self.minimum_offset - offsets[regionui])
        
        #
        self.updateGeometry()
        
        # * Try to center on the new center
        # Function needs to be called after a delay because we need to give geometry
        #   time to adjust.
        AppRoot.timer.after_delay_do(self.recenter, delay=500)
        
        
        # Remember the last 5 alignment targets
        self.last_aligned_to.append(cluster)
        if len(self.last_aligned_to) > 5:
            del self.last_aligned_to[0]
    def recenter(self):
        # This function never gets the precise scrollbar position first try.
        #   It's related tot he width of the self.clustercontainer changing / not updating in time.
        #   I suspect a race condition, but it's just as likely I'm missing something obvious.
        center_position = (self.minimum_offset/self.clustercontainer.size().width())*self.botscroll.maximum()
        print(f"min:{self.minimum_offset}, width:{self.clustercontainer.size().width()}, "
              f"tgtpos:{center_position}, barmax:{self.botscroll.maximum()}, barmin:{self.botscroll.minimum()}")
        self.botscroll.setValue(int(round(center_position)))
    def cbk_export_to_excel(self):
        self.export_into_excel()
    def cbk_code_interact(self):
        code.interact(local=locals())
    def contextMenuEvent(self, event):
        event.accept()
        menu = qtw.QMenu(self)
        
        #ACTION: Restrict by selection
        action = menu.addAction("Restrict by selection")
        action.triggered.connect(self.restrict_shown_data_by_selection)
        
        #ACTION: Export
        action = menu.addAction("Export into Excel")
        action.triggered.connect(self.cbk_export_to_excel)
        
        #ACTION: Save as PNG
        action = menu.addAction("Save as image")
        action.triggered.connect(self.cbk_save_as_image)
        
        #ACTION: Switch to relative size
        if self.display_policy["size_mode"] == "fixed":
            action = menu.addAction("Use proportional CDS sizes")
            action.triggered.connect(self.cbk_use_proportional_widths)
        
        
        #ACTION: Switch to fixed size
        if self.display_policy["size_mode"] == "proportional":
            action = menu.addAction("Use fixed CDS sizes")
            action.triggered.connect(self.cbk_use_fixed_widths)
        
        # Generate menu at click location
        menu.popup(event.globalPos())
    def get_active_clustering(self):
        if self.display_policy["cluster"] == "hetcluster":
            return(self.neighborhood.dataset.heth_clustering)
        elif self.display_policy["cluster"] == "homcluster":
            return(self.neighborhood.dataset.homo_clustering)
        else:
            assert False, "Invalid display policy setting."
    def set_settings(self, settings):
        self.settings = settings
    def restrict_shown_data_by_selection(self):
        dataset = self.neighborhood.dataset
        clusters = set(AppRoot.ui_topclusters.selection)
        regions_to_display = []
        for region in dataset.subset["hits"]:
            clusters_in_region = set()
            for ft in region.fts_borders.values():
                clusters_in_region.add(
                    dataset.heth_clustering.member_to_cluster.get(ft.ref.accession))
            if len(clusters) == len(clusters_in_region.intersection(clusters)):
                regions_to_display.append(region)
        self.display_neighborhoods(regions_to_display=regions_to_display)
    # Export
    def cbk_save_as_image(self):
        if not self.data_loaded:
            msg = qtw.QMessageBox()
            msg.setWindowTitle("Operation Failed")
            msg.setText("Neighborhood view not yet generated.")
            msg.exec_()
            return
        
        path,suffix = qtw.QFileDialog.getSaveFileName(
                        caption="Save output",
                        filter="PNG(*.png);;JPEG(*.jpg *.jpeg)")
        if not path:
            return
        
        headsize = self.headercontainer.size()
        tailsize = self.clustercontainer.size()
        
        # Combined pixmap has the same 
        #   height as either and combined width of both
        pixmapCombo = qtg.QPixmap(headsize.width()+tailsize.width(),
                                  headsize.height())
        #pixmap.set_device_pixel_ratio(2)
        
        
        offset = qtc.QPoint(0,0)
        self.headercontainer.render(pixmapCombo, targetOffset=offset)
        offset = qtc.QPoint(headsize.width(),0)
        self.clustercontainer.render(pixmapCombo, targetOffset=offset)
        
        pixmapCombo.save(path, quality=100)
    def export_into_excel(self):
        if not self.data_loaded:
            msg = qtw.QMessageBox()
            msg.setWindowTitle("Operation Failed")
            msg.setText("Neighborhood view not yet generated.")
            msg.exec_()
            return
        
        self.excel_exporter = ExcelExportFrame(self)
        self.excel_exporter.show()
    def _export_into_excel(self, align=True, unpack=False, separator=", ",
                   cell_variables=None, comment_variables=None,
                   filepath="./excel_output.xlsx", transposed=False):
        wbk = pxl.Workbook()
        
        row = 0
        col = 0
        
        # Used to invert table coordinates if some mad soul wants their table rotated 90°
        #   We're all for user choice here at the Institute of Microbiology
        #   even if said choices are rather unwise.
        def _row():
            return(row if not transposed else col)
        def _col():
            return(col if not transposed else row)
        
        # New variable getters for dynamic variable assignment.
        def getter_proteingroup_allannotations(ft,cl):
            composition = cl._get_composition()
            return("\n".join([f"{key}: {composition[key]}" for key in composition]))
        variable_getters = {
            "feature_position": 
                lambda ft,cl: f"{ft.start}:{ft.stop}",
            "feature_length": 
                lambda ft,cl: f"{ft.stop-ft.start}",
            "protein_accession": 
                lambda ft,cl: ft.ref.accession,
            "protein_annotation": 
                lambda ft,cl: ft.ref.type if ft.ref.type else "N/A",
            "protein_sequence": 
                lambda ft,cl: ft.ref.seq if ft.ref.seq else "N/A",
            "proteingroup_identifier": 
                lambda ft,cl: cl._id,
            "proteingroup_annotation": 
                lambda ft,cl: cl.get_annotation(),
            "proteingroup_allannotations": 
                getter_proteingroup_allannotations,
            "proteingroup_name": 
                lambda ft,cl: cl.get_name(),
            "proteingroup_size": 
                lambda ft,cl: str(len(cl.members)),
            "proteingroup_allaccessions":
                lambda ft,cl: "+".join(cl.members),
        }
        
        #Configure variables to display/defaults
        variables_to_display = {}
        if cell_variables:
            variables_to_display["cell"] = cell_variables
        else:
            variables_to_display["cell"] = ["proteingroup_name", 
                                            "protein_accession"]
        if comment_variables:
            variables_to_display["comment"] = comment_variables
        else:
            variables_to_display["comment"] = ["feature_position", 
                                               "protein_annotation"]
        
        # NOTE: self.active_regionuis may be inappropriate reference
        #       in newer versions, as it's only meant to contain
        #       currently visible rows.
        #       Alternatively, use a third variable to show currently
        #       visisble only :)
        all_seqs = self.displayed_regions
        
        # * * * Set up styles
        
        sh_seqs = wbk.active
        sh_seqs.title = "Sequences"
        sh_table = wbk.create_sheet(title="Overview Table")
        sh_binary = wbk.create_sheet(title="Presence-absence Table")
        
        # * * * Pre-process Sequences
        
        # * First, figure out the offset
        
        base_offset = max([len(reg.fts_borders) for reg in all_seqs])
        
        
        # * Make regex patterns to retrieve colour values from style sheets.
        pattern_background = re.compile(
            r"\sbackground-color:.*?rgb\((?P<color>\d{1,3}.*?\d{1,3}.*?\d{1,3})\)")
        pattern_text = re.compile(
            r"\scolor:.*?rgb\((?P<color>\d{1,3}.*?\d{1,3}.*?\d{1,3})\)")
        
        # * The cluster results to ID the clusters of features
        clustering_results = self.get_active_clustering()
        
        # * * * Write the table
        for ii in range(len(all_seqs)):
            # For each sequence, prepare all information, then write it.
            
            # * First we need to prepare all the information about the sequence
            seq_fts = sorted([ft if ft.type=="cds" else None for ft in all_seqs[ii].fts_borders.values()], key=lambda ft: ft.start)
            for i in reversed(range(0, len(seq_fts))):
                if seq_fts[i] is None:
                    del seq_fts[i]
            seq_clusters = []
            for ft in seq_fts:
                try: 
                    seq_clusters.append(clustering_results.member_to_cluster[ft.ref.accession])
                except KeyError:
                    seq_clusters.append(None)
            
            # Find the align reference
            target = None
            for last_aligned in reversed(self.last_aligned_to):
                if last_aligned in seq_clusters:
                    target = last_aligned
                    break
            
            if target and align:
                offset = base_offset - seq_clusters.index(target) + 0
            else:
                offset = 0
            
            col=ii+1
            
            # * Write the header (with info about organism/sequence)
            # TODO: Getters for txname, txid, contig name
            row = 1
            first_cell = sh_seqs.cell(column=_col(), row=_row(), value=ft.sc.tx.sciname)
            
            row += 1
            sh_seqs.cell(column=_col(), row=_row(), value=ft.sc.tx.taxid)
            
            row += 1
            sh_seqs.cell(column=_col(), row=_row(), value=ft.sc.accession)
            
            row += 1
            startstops = []
            for ft in seq_fts:
                startstops.append(ft.start)
                startstops.append(ft.stop)
            position = f"{min(startstops)}:{max(startstops)}"
            sh_seqs.cell(column=_col(), row=_row(), value=position)
            
            header_row = row
            row += 1
            # Apply offset
            row = row + offset
            # * Now we write each feature in turn
            for i in range(len(seq_fts)):
                row = row + 1
                
                # * Style
                #   Get the colours from the stylesheet and conver to hexadecimal.
                #   The string expression ("%02x%02x%02x" % tuple) accepts a tuple
                #   3 ints corresponding to Red, Green and Blue intensities.
                cluster = seq_clusters[i]
                feature = seq_fts[i]
                
                if seq_clusters[i]:
                    # Value and comments are annoyingly large expressions.
                    #   The expression iterates over a list of variable names,
                    #   retrieves corresponding getter functions from
                    #   variable_getters, calls them to retrieve the values
                    #   and then joins them with a given separator.
                    txt_color = "%02x%02x%02x" % tuple([int(x) for x in re.search(pattern_text, seq_clusters[i].active_style)["color"].split(",")])
                    bg_color = "%02x%02x%02x" % tuple([int(x) for x in re.search(pattern_background, seq_clusters[i].active_style)["color"].split(",")])
                    value = separator.join([variable_getters[varname](feature,cluster) for varname in variables_to_display["cell"]])
                    comment = "\n".join([variable_getters[varname](feature,cluster) for varname in variables_to_display["comment"]])
                else:
                    txt_color = "%02x%02x%02x" % (255,255,255)
                    bg_color = "%02x%02x%02x" % (0,0,0)
                    value = "N/A"
                    comment = seq_fts[i].ref.accession
                # * Comment
                cell = sh_seqs.cell(column=_col(), row=_row(), value=value)
                cell.comment = pxl.comments.Comment(comment, "")
                cell.fill = pxl.styles.PatternFill("solid", fgColor=bg_color)
                cell.font = pxl.styles.Font(color=txt_color)
        
        # Set fixed width for all columns EDIT: Breaks the sheet, no clue why
        #for column in sh_seqs.iter_cols(min_col=1, max_col=sh_seqs.max_column):
        #    sh_seqs.column_dimensions[column] = 30
        
        # Freeze the header pane
        row = header_row+1
        col = 1
        topleft = sh_seqs.cell(column=_col(), row=_row())
        sh_seqs.freeze_panes = topleft
        
        # * * * Write Overview Table
        
        # col and row values on the second sheet are *not* invertable
        
        # Get overview table
        otable = AppRoot.ui_topclusters.cluster_table
        # Copy headers
        for col in range(0, otable.columnCount()):
            item = otable.horizontalHeaderItem(col)
            sh_table.cell(column=col+1, row=1, value=item.text())
        # Copy cells
        for row in range(0, otable.rowCount()):
            for col in range(0, otable.columnCount()):
                item = otable.item(row, col)
                if item is None:
                    # If it's not an item, we can assume it's
                    #   a ColourableButton until more widgets are inserted
                    item = otable.cellWidget(row, col)
                    value = item.get_active_cluster().get_name()
                    cell = sh_table.cell(column=col+1, row=row+2, value=value)
                    
                    cluster_stylesheet = item.get_active_cluster().active_style
                    txt_color = "%02x%02x%02x" % tuple([int(x) for x in re.search(pattern_text, cluster_stylesheet)["color"].split(",")])
                    bg_color = "%02x%02x%02x" % tuple([int(x) for x in re.search(pattern_background, cluster_stylesheet)["color"].split(",")])
                    cell.font = pxl.styles.Font(color=txt_color)
                    cell.fill = pxl.styles.PatternFill("solid", fgColor=bg_color)
                else:
                    value = item.text()
                    cell = sh_table.cell(column=col+1, row=row+2, value=value)
        
        # * * * Write Binary Table
        # It's like a blend between overview table and neighborhood table!
        
        # First, get a dictionary of positions in the table
        #   Mostly because I trust a hash table more than the list.index()
        #   lookup function.
        all_seqs_indices = {}
        for i in range(len(all_seqs)):
            all_seqs_indices[all_seqs[i].sc] = i
        
        #The column at which values start (and before which headers end)
        startcol = 3
        for ii in range(len(all_seqs)):
            # * Write the header (with info about organism/sequence)
            # TODO: Getters for txname, txid, contig name
            row = 1
            first_cell = sh_binary.cell(column=startcol+ii, row=row, 
                                      value=all_seqs[ii].sc.tx.sciname)
            
            row += 1
            sh_binary.cell(column=startcol+ii, row=row, 
                         value=all_seqs[ii].sc.tx.taxid)
            
            row += 1
            sh_binary.cell(column=startcol+ii, row=row, 
                         value=all_seqs[ii].sc.accession)
            
            row += 1
            startstops = []
            for ft in all_seqs[ii].fts_borders.values():
                startstops.append(ft.start)
                startstops.append(ft.stop)
            position = f"{min(startstops)}:{max(startstops)}"
            sh_binary.cell(column=startcol+ii, row=row, value=position)
        
        #The row at which values start (and before which headers end)
        startrow = row+1
        
        # Now generate the protein group rows and populate them
        for tablerow in range(0, otable.rowCount()):
            cluster = otable.cellWidget(tablerow, 1).get_active_cluster()
            annotation = otable.item(tablerow, 2).text()
            
            # Write header
            sh_binary.cell(column=1, row=tablerow+startrow, 
                           value=cluster.get_name())
            sh_binary.cell(column=2, row=tablerow+startrow,
                           value=annotation)
            
            # Fill out presence/absence data
            for member in cluster.members.values():
                for feature in member.fts_borders.values():
                    sc = feature.sc
                    if sc in all_seqs_indices:
                        # Expressions for colors etc. copied from above
                        txt_color = "%02x%02x%02x" % tuple([int(x) for x in re.search(pattern_text, cluster.active_style)["color"].split(",")])
                        bg_color = "%02x%02x%02x" % tuple([int(x) for x in re.search(pattern_background, cluster.active_style)["color"].split(",")])
                        value = separator.join([variable_getters[varname](feature,cluster) for varname in variables_to_display["cell"]])
                        comment = "\n".join([variable_getters[varname](feature,cluster) for varname in variables_to_display["comment"]])
                        
                        region_index = all_seqs_indices[sc]
                        cell = sh_binary.cell(column=startcol+region_index, 
                                            row=tablerow+startrow, 
                                            value=value)
                        cell.comment = pxl.comments.Comment(comment, "")
                        cell.fill = pxl.styles.PatternFill("solid", 
                                                           fgColor=bg_color)
                        cell.font = pxl.styles.Font(color=txt_color)
                
        wbk.save(filename=filepath)
class NeighborhoodViewer(qtw.QWidget):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        #self.btns_added = None
        #self.dummies = None #debug
        self.visible = set() #debug
        self.lastupdate = 0
        
        #Scrollbar hack
        self.firstshow = True
        self.botscroll = None #The bottom scrollbar
        
        #
        self.dataset = None
        self.clusters = None
        
        self.view = NeighborhoodView(self)
        
        #
        self.nsize = 20 #displayed neighborhood size in either direction
        
        self._init_ui()
    def _init_ui(self):
        self.mylayout = qtw.QHBoxLayout()
        self.setLayout(self.mylayout)
        self.mylayout.setContentsMargins(0,0,0,0)
        self.mylayout.addWidget(self.view)
    
    def prepare_neighborhood(self, settings):
        # Explanation:
        #   This function is a bit difficult to keep track of.
        #   Step-1: Download data from NCBI
        #   Step-2: Protein clustering
        #   Step-3: Display neighborhood
        #
        #   Steps 1 and 2 are executed in a work thread.
        #   Each subsequent step is executed via a callback function
        #   triggered at the end of a given work thread.
        #   For this reason, the steps are prepared in reverse order,
        #   and the sequence is then triggered with the first.
        
        # First, we find our dataset
        self.dataset = AppRoot.ui_constraintsframe.dataset
        
        # and create a progress dialog for steps 2 and 3
        AppRoot.progress_dialog.reinitialize(
                              title="Neighborhood View Progress", 
                              label_text="Clustering", 
                              on_abort=None,
                              text_only=True)
        
        # PREPARE STEP-3: DISPLAY NEIGHBORHOOD
        def on_clustering_finished():
            # TODO: This step is performed in main thread as it modifies UI
            #   Some lag will be inevitable until some painful
            #   refactoring is finally made happen.
            self.view.set_settings(settings)
            self.view.display_neighborhoods()
        
            # Set the cluster table
            AppRoot.ui_topclusters.set_neighborhood(self.view)
            AppRoot.progress_dialog.setHidden(True)
        
        
        # PREPARE STEP-2: CLUSTERING
        worker_clustering = WorkerThread(
          self._prepare_neighborhood, settings)
        worker_clustering.signals.finished.connect(on_clustering_finished)
        
        def on_download_finished():
            AppRoot.timer.after_delay_do(worker_clustering.safe_start)
        
        
        # PREPARE STEP-1: NCBI DOWNLOAD
        worker_download = WorkerThread(
          self.dataset.get_neighborhoods_for_subset)
        worker_download.signals.finished.connect(on_download_finished)
        
        # With this somewhat confusing implementation, the
        #   download manager calls the signal in one thread,
        #   then responds to it in the other.
        worker_download.signals.progress_manager_update.connect(
            dbdl.DOWNLOAD_MANAGER.cbk_status_update)
        dbdl.DOWNLOAD_MANAGER.set_progress_manager_update_signal(
            worker_download.signals.progress_manager_update)
        
        # EXECUTE STEP-1
        
        AppRoot.progress_dialog.show()
        # * Get neighborhoods from NCBI
        print("Obtaining neighbourhoods...\n")
        worker_download.safe_start()
    def _prepare_neighborhood(self, settings):
        # * Cluster proteins in neighborhood
        #First we need to make a list of all proteins we want processed
        #TODO:
        #NOTE: The order in which proteins are fed to the clustering
        #      algorithm CAN affect the results.
        #      This may be an issue for consistency / reproducibility.
        #      Do we want to be sorting them here? Do we have how?
        
        print("Determining aminoacid sequences for clustering...")
        
        proteins_to_cluster = set()
        for gr in self.dataset.subset["hits"]:
            for ft in gr.fts_borders.values():
                if ft.type == "cds":
                    proteins_to_cluster.add(ft.ref.accession)
        
        print(f"Prepped {len(proteins_to_cluster)} proteins for clustering.\n")
        print("    Beginning clustering...\n")
        
        #Finally, we run USEARCH (hopefully it's in the process folder)
        AppRoot.cluster_manager.fasta_dump(proteins_to_cluster, 
                                           self.dataset.root)
        
        
        print(f"    Clustering to identity {settings['global_clustering_identity_threshold']}")
        AppRoot.cluster_manager.run_usearch(\
            identity=settings["global_clustering_identity_threshold"])
        
        
        self.dataset.homo_clustering = AppRoot.cluster_manager.parse_cluster_results(
            self.dataset.root,
            identity=settings["global_clustering_identity_threshold"])
            
        
        self.dataset.heth_clustering = AppRoot.cluster_manager.ublast_meta_cluster(
            clustering_result=self.dataset.homo_clustering,
            root=self.dataset.root, 
            evalue=settings["local_clustering_evalue_threshold"],
            identity=settings["local_clustering_identity_threshold"])
        
    def find_homologs(self, query_cluster):
        #TODO: This function needs a rewrite.
        AppRoot.cluster_manager.p_blastp = "blastp"
        
        # First, we get the cluster centroids into a fasta
        centroid_accessions = []
        # TODO: Specifically, self.clustering needs to be replaced
        #       by a switch for self.dataset.heth/homclustering
        for cluster in self.clustering.clusters.values():
            centroid_accessions.append(cluster.centroid.accession)
        AppRoot.cluster_manager.fasta_dump(centroid_accessions,
                                          self.dataset.root)
        
        # Then we run blastp
        print("Running local ublast search...")
        align_results = AppRoot.cluster_manager.neighborhood_ublast(
            query_cluster.centroid.seq)
        #Convert it to a list of values.
        align_results = [x for x in align_results.values()]
        print("Done.")
        
        # Finally, we generate the UI
        widget = qtw.QWidget()
        layout = qtw.QVBoxLayout()
        
        print(align_results)
        
        for alignment in sorted(align_results, reverse=False,
                key=lambda alignment: alignment["evalue"]):
            cluster = self.clustering.\
                member_to_cluster[alignment["hit"]]
            
            subwidget = qtw.QWidget()
            sublayout = qtw.QHBoxLayout()
            
            if alignment["evalue"] == 0:
                evalue = 0
            else:
                evalue = round(math.log(alignment["evalue"], 10), 0)
            
            sublayout.addWidget(cluster.get_clus_button(nofeature=True))
            sublayout.addWidget(qtw.QLabel(f"evalue:{evalue}"))
            
            subwidget.setLayout(sublayout)
            
            layout.addWidget(subwidget)
        
        widget.setLayout(layout)
        
        print("Showing local blast search...")
        widget.show()
        
        #Hook the widget
        AppRoot.homolog_widget = widget

class TopClusters(qtw.QWidget, Ui_TopClusters):
    class SelectionCheckbox(qtw.QTableWidgetItem):
        def __init__(self, topclusters, cluster, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsUserCheckable
                          | qtc.Qt.ItemIsEnabled)
            self.topclusters = topclusters
            self.cluster = cluster
            
            
            self.setCheckState(qtc.Qt.Unchecked)
            
            # This is an absolutely disgusting hack to make
            #   utilizing whitespace to sort TableWidgetItems
            #   already containing something else
            self.setData(qtc.Qt.DisplayRole, "")
            self.setText(":(")
        def cbk_selection_change(self):
            if self.checkState() is qtc.Qt.Checked:
                self.topclusters.select_cluster(self.cluster)
                self.setText("✓")
            elif self.checkState() is qtc.Qt.Unchecked:
                self.topclusters.unselect_cluster(self.cluster)
                self.setText(" ")
        def check(self):
            self.setCheckState(qtc.Qt.Checked)
            # See init
            #self.setData(qtc.Qt.DisplayRole, ":)")
        def uncheck(self):
            self.setCheckState(qtc.Qt.Unchecked)
            # See init
            #self.setData(qtc.Qt.DisplayRole, ":(")
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.init_ui()
        
        self.cluster_rows = []
        
        self.selection = {}
        self.selection_checkboxes = {}
        
        self.cluster_table.itemChanged.connect(self.cbk_item_changed)
    def init_ui(self):
        self.setupUi(self)
        
        self.clear_btn.clicked.connect(self.cbk_clear_selection)
        
    def cbk_item_changed(self, item):
        if isinstance(item, self.SelectionCheckbox):
            item.cbk_selection_change()
    def cbk_clear_selection(self):
        for cluster in list(self.selection):
            self.unselect_cluster(cluster)
    def set_neighborhood(self, neiview):
        print("DEBUG: Creating table.")
        #TODO: Run this again when displaypolicy switches from homclusters to hetclusters
        self.neiview = neiview
        
        clustering = self.neiview.get_active_clustering()
        cluster_instances_unique = self.neiview.cluster_instances_unique
        cluster_instances_total = self.neiview.cluster_instances_total
        
        #Survey all displayed neighbourhoods/gclusters
        #  to get some per-gcluster statistics
        
        
        
        
        # * * SET UP TABLE
        #TODO: Delete all existing elements
        self.cluster_table.clear()
        
        #Set size and headers
        self.cluster_table.setRowCount(len(clustering.clusters))
        self.cluster_table.setColumnCount(7)
        
        self.cluster_table.setHorizontalHeaderLabels(
            ["", "ID", "Annotation", "Unique\nCounts", 
            "Total\nCounts", "Duplication\nCounts", "Marker"])
        # * * Populate
        #First disable sorting so it doesn't mess with us during population
        self.cluster_table.setSortingEnabled(False)
        
        # Set the width of 0th column to 20
        # (Doesn't work right now)
        sh = self.cluster_table.horizontalHeaderItem(0).sizeHint()
        sh.setWidth(20)
        self.cluster_table.horizontalHeaderItem(0).setSizeHint(sh)
        
        #Do the thing
        row = 0
        for cluster in clustering.clusters.values():
            #Cluster ID
            #item = qtw.QTableWidgetItem()
            #item.setText(str(cluster._id))
            #item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            #self.cluster_table.setItem(row, 0, item)
            
            #COLUMN 0: Checkmark
            item = self.SelectionCheckbox(self, cluster)
            self.cluster_table.setItem(row, 0, item)
            self.selection_checkboxes[cluster] = item
                                   
            #COLUMN 1: Button
            #This is a rather ugly hack, but.
            btn = ColorableButton(homcluster=cluster, hetcluster=cluster)
            self.cluster_table.setIndexWidget(self.cluster_table.model().index(row, 1), 
                                              btn)
            
            #COLUMN 2: Annotation
            # TODO: The annotations don't update when the user makes a change -_-
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, cluster.get_annotation())
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.cluster_table.setItem(row, 2, item)
            
            #COLUMN 3: Unique Counts
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, cluster_instances_unique[cluster])
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.cluster_table.setItem(row, 3, item)
            
            #COLUMN 4: Total Counts
            total = sum([x*cluster_instances_total[cluster][x] for x in cluster_instances_total[cluster]])
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, total)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.cluster_table.setItem(row, 4, item)
            
            #COLUMN 5: Duplication
            instances = ", ".join([f"{count}:{cluster_instances_total[cluster][count]}"\
                for count in sorted(cluster_instances_total[cluster])[1:]])
            
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, instances)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.cluster_table.setItem(row, 5, item)
            
            #COLUMN 6: Marker?
            value = "Yes" if hasattr(cluster, "_temp_is_marker") else "No"
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, value)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.cluster_table.setItem(row, 6, item)
            
            row += 1
        
        #Re-enable sorting
        self.cluster_table.setSortingEnabled(True)
    def select_cluster(self, cluster):
        # If it's already selected, ignore the call
        if cluster in self.selection: return
        
        # Otherwise:
        # Add colourable button of the right cluster to selection
        btn = ColorableButton(homcluster=cluster, hetcluster=cluster)
        self.selected_lay.addWidget(btn)
        
        # Remember the button
        self.selection[cluster] = btn
        
        # Check the checkbox, just in case the selection
        #   command came from somewhere other than checkbox.
        self.selection_checkboxes[cluster].setCheckState(qtc.Qt.Checked)
    def unselect_cluster(self, cluster):
        # If it's not actually selected, ignore the call
        if cluster not in self.selection: return
        
        # Otherwise:
        # Remove colourable button from selection layout
        self.selection[cluster].setParent(None)
        #self.selected_lay.removeWidget(self.selection[cluster])
        
        # Forget the button
        del self.selection[cluster]
        
        #Uncheck the checkbox
        self.selection_checkboxes[cluster].setCheckState(qtc.Qt.Unchecked)
    def contextMenuEvent(self, event):
        event.accept()
        menu = qtw.QMenu(self)
        
        #ACTION: Restrict by selection
        action = menu.addAction("Restrict by selection")
        action.triggered.connect(AppRoot.\
            ui_neighborhoodviewer.view.restrict_shown_data_by_selection)
        
        # Generate menu at click location
        menu.popup(event.globalPos())

class SelectorWidget(qtw.QWidget, Ui_SelectorWidget):
    def __init__(self, name_a, name_b, items_a=[], items_b=[], *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.init_ui(name_a, name_b, items_a, items_b)
    def init_ui(self, name_a, name_b, items_a, items_b):
        self.setupUi(self)
        
        self.ui_listA_label.setText(name_a)
        self.ui_listB_label.setText(name_b)
        
        self.ui_btn_AtoB.clicked.connect(self.cbk_atob)
        self.ui_btn_BtoA.clicked.connect(self.cbk_btoa)
        self.ui_btn_Bup.clicked.connect(self.cbk_bup)
        self.ui_btn_Bdown.clicked.connect(self.cbk_bdown)
        
        for item in items_a:
            self.add_a_item(item)
        for item in items_b:
            self.add_b_item(item)
    def cbk_atob(self):
        # An overly elaborate one liner.
        # currentItem() returns an object,
        #   but takeItem needs row
        self.ui_listB.addItem( # Move item
            self.ui_listA.takeItem( # Get item from row and remove from A
                self.ui_listA.row( # Get row of current item
                    self.ui_listA.currentItem() # Get current item
                    )))
    def cbk_btoa(self):
        # An overly elaborate one liner.
        # currentItem() returns an object,
        #   but takeItem needs row
        self.ui_listA.addItem( # Move item
            self.ui_listB.takeItem( # Get item from row and remove from B
                self.ui_listB.row( # Get row of current item
                    self.ui_listB.currentItem() # Get current item
                    )))
        self.ui_listA.sortItems()
    def cbk_bup(self):
        # Take item from list, put it one higher.
        row = self.ui_listB.row(self.ui_listB.currentItem())
        if row==0: return #Don't move if it's at the top
        item = self.ui_listB.takeItem(row)
        self.ui_listB.insertItem(row-1, item)
        self.ui_listB.setCurrentRow(row-1)
    def cbk_bdown(self):
        # Take item from list, put it one lower.
        
        #Get Row
        row = self.ui_listB.row(self.ui_listB.currentItem())
        
        #Don't move down if there's nothing lower
        if not self.ui_listB.item(row+1): 
            print("F")
            return
        item = self.ui_listB.takeItem(row)
        self.ui_listB.insertItem(row+1, item)
        self.ui_listB.setCurrentRow(row+1)
    def get_a_contents(self):
        contents = []
        for i in range(self.ui_listA.count()):
            contents.append(self.ui_listA.item(i).text())
        return(contents)
    def get_b_contents(self):
        contents = []
        for i in range(self.ui_listB.count()):
            contents.append(self.ui_listB.item(i).text())
        return(contents)
    
    #def get_a_contents(self):
    
    #def get_b_contents(self):
    
    def add_a_item(self, label):
        self.ui_listA.addItem(label)
    def add_b_item(self, label):
        self.ui_listB.addItem(label)

class ExcelExportFrame(qtw.QDialog, Ui_ExcelExportFrame):
    # Possible variables to be displayed in cells within the excel sheet.
    # These values should be the same as the property getters used in
    #   NeighborhoodView's into_excel() function.
    # TODO: Put the dictionary somewhere shared, so we're not running two
    #   versions of it that need to be synchronized.
    cell_content = [
        "proteingroup_allaccessions",
        "proteingroup_size",
        "proteingroup_name",
        "proteingroup_allannotations",
        "proteingroup_annotation",
        "proteingroup_identifier",
        "protein_sequence",
        "protein_annotation",
        "protein_accession",
        "feature_length",
        "feature_position",
    ]
    cell_defaults = ["proteingroup_name", "protein_accession"]
    comment_defaults = ["feature_position", "protein_annotation"]
    def __init__(self, neighborhood_view):
        super().__init__()
        self.init_ui()
        
        self.neighborhood_view = neighborhood_view
    def init_ui(self):
        self.setupUi(self)
        self.setModal(True)
        
        # cell content selector
        self.cell_contentselector = SelectorWidget(
            "Possible Values", 
            "Values to Display in Cells")
        self.selector_container_cell_lay.addWidget(self.cell_contentselector)
        
        # comment content selector
        self.comment_contentselector = SelectorWidget(
            "Possible Values", 
            "Values to Display in Comments")
        self.selector_container_comment_lay.addWidget(self.comment_contentselector)
        
        # load variables into both selectors
        for item in self.cell_content:
            if item not in self.cell_defaults:
                self.cell_contentselector.add_a_item(item)
            else:
                self.cell_contentselector.add_b_item(item)
            if item not in self.comment_defaults:
                self.comment_contentselector.add_a_item(item)
            else:
                self.comment_contentselector.add_b_item(item)
        
        #TODO: Set title to "Export Data into Excel"
        
        self.io_save_btn.clicked.connect(self.cbk_accept)
        self.io_cancel_btn.clicked.connect(self.cbk_cancel)
    
    def obtain_settings(self, path):
        settings = {}
        settings["align"] = self.align_chk.isChecked()
        settings["unpack"] = False
        settings["separator"] = self.cell_value_separator_ledit.text()
        settings["cell_variables"] = self.cell_contentselector.get_b_contents()
        settings["comment_variables"] = self.comment_contentselector.get_b_contents()
        settings["filepath"] = path if path else "./excel_output.xlsx"
        settings["transposed"] = self.transpose_table_chk.isChecked()
        return(settings)
    def generate_excel(self):
        path,filter = qtw.QFileDialog.getSaveFileName(
                        caption="Save output",
                        filter="Excel 2010 Spreadsheet (*.xlsx)")
        if not path:
            return
        settings = self.obtain_settings(path)
        self.neighborhood_view._export_into_excel(**settings)
        self._close()
    def cbk_accept(self):
        self.generate_excel()
    def cbk_cancel(self):
        self._close()
    def _close(self):
        self.setHidden(True)
        self.neighborhood_view.excel_exporter = None
    
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

 #                            Other                                  #

 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def launch_app():
    app = AppRoot()
    app.launch()
